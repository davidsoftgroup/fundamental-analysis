import streamlit as st
import sys
import os
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from utils.database import get_connection, init_db

st.set_page_config(page_title="مقایسه شرکت‌ها", layout="wide")
from utils.styles import apply_styles
apply_styles()


st.markdown("""
<style>
    .main .block-container { direction: rtl; padding-top: 1.2rem; }
    .stTable, table { direction: rtl; text-align: center; }
    h1 { font-size: 1.6rem !important; }
    h2, h3 { font-size: 1.2rem !important; }
</style>
""", unsafe_allow_html=True)

st.title("مقایسه شرکت‌ها")

init_db()

COLORS = ["#2563eb", "#dc2626", "#16a34a", "#ea580c", "#7c3aed", "#0891b2", "#db2777", "#0d9488"]

def get_companies():
    conn = get_connection()
    rows = conn.execute("""
        SELECT id, symbol, name_fa, industry, market_value, rank_in_industry,
               fiscal_end_month, fiscal_end_day
        FROM companies ORDER BY symbol
    """).fetchall()
    conn.close()
    return rows

def get_periods(company_id):
    conn = get_connection()
    rows = conn.execute("""
        SELECT p.id, p.year_solar, p.period_type, p.end_month, p.end_day,
               f.operating_revenue, f.cogs, f.other_income, f.non_operating_income,
               f.net_profit, f.comprehensive_income, f.inventory,
               f.trade_receivables, f.equity, f.current_assets,
               f.total_assets, f.approved_dividend
        FROM periods p
        LEFT JOIN financials f ON f.period_id = p.id
        WHERE p.company_id = ?
        ORDER BY p.year_solar DESC, p.period_type DESC
    """, (company_id,)).fetchall()
    conn.close()
    return rows

def get_all_monthly_sales(company_id):
    conn = get_connection()
    rows = conn.execute("""
        SELECT year_solar, month, domestic_sales, export_sales, total_sales
        FROM monthly_sales WHERE company_id = ?
        ORDER BY year_solar, month
    """, (company_id,)).fetchall()
    conn.close()
    return rows

def get_recurring_by_year(company_id):
    conn = get_connection()
    rows = conn.execute("""
        SELECT year_solar, SUM(amount) as total
        FROM non_operating_items
        WHERE company_id = ? AND is_recurring = 1
        GROUP BY year_solar ORDER BY year_solar
    """, (company_id,)).fetchall()
    conn.close()
    return {r["year_solar"]: float(r["total"] or 0) for r in rows}

def safe_div(a, b):
    try:
        if a is None or b is None or float(b) == 0:
            return None
        return float(a) / float(b)
    except Exception:
        return None

def calc_metrics(row):
    revenue = row["operating_revenue"]
    other_inc = row["other_income"] or 0
    non_op = row["non_operating_income"] or 0
    net_profit = row["net_profit"]
    receivables = row["trade_receivables"]
    current_assets = row["current_assets"]
    comprehensive = row["comprehensive_income"]
    dividend = row["approved_dividend"]

    # فقط درآمدهای مثبت از سود خالص کم می‌شوند
    # اگر منفی بود (هزینه) → صفر در نظر گرفته می‌شود
    other_adj = float(other_inc) if other_inc and float(other_inc) > 0 else 0
    non_op_adj = float(non_op) if non_op and float(non_op) > 0 else 0

    op_profit = None
    if net_profit is not None:
        op_profit = float(net_profit) - other_adj - non_op_adj

    net_margin = safe_div(op_profit, revenue)
    recv_ratio = safe_div(receivables, current_assets)
    div_ratio = safe_div(dividend, comprehensive)
    return op_profit, net_margin, recv_ratio, div_ratio
    
    

def get_first_quarter_months(fiscal_end_month):
    start = (fiscal_end_month % 12) + 1
    return [((start - 1 + i) % 12) + 1 for i in range(3)]

def calc_forward_estimates(company, periods, monthly_rows):
    result = {
        "has_data": False, "target_year": None,
        "sales_final": None, "margin_avg": None,
        "est_profit": None, "est_recurring_non_op": None,
        "est_dividend": None, "payout_avg": None,
        "pe_forward": None, "pd_forward": None, "ps_forward": None,
        "margin_annual": None, "margin_last": None,
    }
    if not periods and not monthly_rows:
        return result

    fiscal_end_month = company["fiscal_end_month"] or 12
    market_value = company["market_value"]

    years_in_periods = [r["year_solar"] for r in periods] if periods else []
    years_in_monthly = [r["year_solar"] for r in monthly_rows] if monthly_rows else []
    all_years = sorted(set(years_in_periods + years_in_monthly))
    if not all_years:
        return result

    last_year = max(all_years)
    has_annual_last = any(r["year_solar"] == last_year and r["period_type"] == 12 for r in periods)
    target_year = last_year + 1 if has_annual_last else last_year
    result["target_year"] = target_year

    q1 = None
    for r in periods:
        if r["year_solar"] == target_year and r["period_type"] == 3:
            q1 = r
            break
    q1_months = get_first_quarter_months(fiscal_end_month)
    sales_m1 = sales_m2 = None

    if q1 and q1["operating_revenue"] and q1["operating_revenue"] > 0:
        rev = float(q1["operating_revenue"])
        sales_m1 = (rev / 3) * 12
        month_sales = {}
        for r in monthly_rows:
            if r["year_solar"] == target_year and r["month"] in q1_months:
                tot = r["total_sales"]
                if tot is None:
                    tot = (r["domestic_sales"] or 0) + (r["export_sales"] or 0)
                if tot and tot > 0:
                    month_sales[r["month"]] = float(tot)
        if month_sales:
            last_val = month_sales[max(month_sales.keys())]
            sum_existing = sum(month_sales.values())
            sales_m2 = (last_val * (12 - len(month_sales))) + sum_existing
        else:
            sales_m2 = sales_m1
    else:
        month_sales = {}
        for r in monthly_rows:
            if r["year_solar"] == target_year and r["month"] in q1_months:
                tot = r["total_sales"]
                if tot is None:
                    tot = (r["domestic_sales"] or 0) + (r["export_sales"] or 0)
                if tot and tot > 0:
                    month_sales[r["month"]] = float(tot)
        if not month_sales:
            return result
        n = len(month_sales)
        sales_m1 = (sum(month_sales.values()) / n) * 12
        last_val = month_sales[max(month_sales.keys())]
        sum_existing = sum(month_sales.values())
        sales_m2 = (last_val * (12 - n)) + sum_existing

    sales_final = min(sales_m1, sales_m2) if sales_m1 and sales_m2 else (sales_m1 or sales_m2)
    result["sales_final"] = sales_final

    margin_annual = margin_last = None
    for r in periods:
        if r["year_solar"] == target_year - 1 and r["period_type"] == 12:
            _, margin, _, _ = calc_metrics(r)
            if margin is not None:
                margin_annual = margin
            break
    candidates = [r for r in periods if r["year_solar"] <= target_year]
    if candidates:
        non_annual = [r for r in candidates if r["period_type"] != 12]
        last_period = max(non_annual or candidates, key=lambda x: (x["year_solar"], x["period_type"]))
        _, margin, _, _ = calc_metrics(last_period)
        if margin is not None:
            margin_last = margin

    result["margin_annual"] = margin_annual
    result["margin_last"] = margin_last
    if margin_annual is not None and margin_last is not None:
        margin_avg = (margin_annual + margin_last) / 2
    else:
        margin_avg = margin_annual if margin_annual is not None else margin_last
    if margin_avg is None or sales_final is None:
        return result
    result["margin_avg"] = margin_avg

    est_profit = sales_final * margin_avg
    recurring_map = get_recurring_by_year(company["id"])
    relevant = [y for y in sorted(recurring_map.keys()) if y < target_year]
    est_recurring = None
    if len(relevant) >= 2:
        v1, v2 = recurring_map[relevant[-2]], recurring_map[relevant[-1]]
        g = (v2 - v1) / v1 if v1 != 0 else 0
        est_recurring = v2 * (1 + g)
    elif len(relevant) == 1:
        est_recurring = recurring_map[relevant[-1]]
    result["est_recurring_non_op"] = est_recurring
    if est_recurring is not None:
        est_profit = est_profit + est_recurring
    result["est_profit"] = est_profit

    payouts = []
    for y in [target_year - 1, target_year - 2]:
        for r in periods:
            if r["year_solar"] == y and r["period_type"] == 12:
                div, comp = r["approved_dividend"], r["comprehensive_income"]
                if div is not None and comp is not None and float(comp) != 0:
                    payouts.append(float(div) / float(comp))
                break
    payout_avg = sum(payouts) / len(payouts) if payouts else None
    result["payout_avg"] = payout_avg
    if payout_avg is not None and est_profit is not None:
        result["est_dividend"] = est_profit * payout_avg

    if market_value and est_profit and est_profit > 0:
        result["pe_forward"] = market_value / est_profit
    if market_value and result.get("est_dividend") and result["est_dividend"] > 0:
        result["pd_forward"] = market_value / result["est_dividend"]
    if market_value and sales_final and sales_final > 0:
        result["ps_forward"] = market_value / sales_final

    result["has_data"] = True
    return result

def fmt(v):
    if v is None:
        return "—"
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return "—"

def fmt_pct(v):
    if v is None:
        return "—"
    try:
        return f"{float(v)*100:.1f}%"
    except Exception:
        return "—"

def fmt_ratio(v):
    if v is None:
        return "—"
    try:
        return f"{float(v):.2f}"
    except Exception:
        return "—"

def pe_color(pe):
    """رنگ شرطی P/E"""
    if pe is None:
        return "#6b7280"
    if pe < 5:
        return "#16a34a"   # سبز
    if pe <= 7:
        return "#ca8a04"   # زرد
    return "#dc2626"       # قرمز

def pe_label(pe):
    if pe is None:
        return "—"
    if pe < 5:
        return f"{pe:.2f} (مناسب)"
    if pe <= 7:
        return f"{pe:.2f} (متوسط)"
    return f"{pe:.2f} (بالا)"

def pe_html(pe):
    """نمایش رنگی P/E در HTML"""
    if pe is None:
        return "—"
    color = pe_color(pe)
    label = pe_label(pe)
    return f'<span style="color:{color}; font-weight:bold;">{label}</span>'

# ============================================================
# انتخاب شرکت‌ها
# ============================================================
companies = get_companies()
if not companies:
    st.warning("هنوز شرکتی ثبت نشده است.")
    st.stop()

# صنایع موجود
industries = sorted(set(c["industry"] for c in companies if c["industry"]))

col_f1, col_f2 = st.columns(2)
with col_f1:
    mode = st.radio(
        "نحوه انتخاب:",
        options=["انتخاب دستی", "همه شرکت‌ها", "بر اساس صنعت"],
        horizontal=True
    )
with col_f2:
    selected_industry = None
    if mode == "بر اساس صنعت":
        if industries:
            selected_industry = st.selectbox("صنعت:", options=industries)
        else:
            st.warning("صنعتی ثبت نشده است.")
            st.stop()

options = {f"{c['symbol']} — {c['name_fa'] or ''}": c for c in companies}

if mode == "همه شرکت‌ها":
    selected_companies = list(companies)
elif mode == "بر اساس صنعت":
    selected_companies = [c for c in companies if c["industry"] == selected_industry]
    if not selected_companies:
        st.warning("شرکتی در این صنعت نیست.")
        st.stop()
else:
    selected_labels = st.multiselect(
        "شرکت‌های مورد نظر (حداقل ۲):",
        options=list(options.keys()),
        default=list(options.keys())[: min(3, len(options))]
    )
    if len(selected_labels) < 1:
        st.info("حداقل یک شرکت انتخاب کنید.")
        st.stop()
    selected_companies = [options[lb] for lb in selected_labels]

st.caption(f"تعداد شرکت‌های انتخاب‌شده: **{len(selected_companies)}**")

# ============================================================
# محاسبه داده
# ============================================================
rows_data = []
for comp in selected_companies:
    periods = get_periods(comp["id"])
    monthly = get_all_monthly_sales(comp["id"])
    fwd = calc_forward_estimates(comp, periods, monthly)
    latest_margin = latest_recv = None
    if periods:
        _, latest_margin, latest_recv, _ = calc_metrics(periods[0])
    rows_data.append({
        "نماد": comp["symbol"],
        "نام": comp["name_fa"] or "—",
        "صنعت": comp["industry"] or "—",
        "ارزش بازار": comp["market_value"],
        "برآورد فروش": fwd.get("sales_final"),
        "میانگین حاشیه": fwd.get("margin_avg"),
        "برآورد سود": fwd.get("est_profit"),
        "برآورد سود تقسیمی": fwd.get("est_dividend"),
        "P/E Forward": fwd.get("pe_forward"),
        "P/D Forward": fwd.get("pd_forward"),
        "P/S Forward": fwd.get("ps_forward"),
        "حاشیه آخرین دوره": latest_margin,
        "نسبت مطالبات": latest_recv,
    })

df = pd.DataFrame(rows_data)

# ============================================================
# مرتب‌سازی
# ============================================================
# ============================================================
# جدول مقایسه + مرتب‌سازی تمیز
# ============================================================
st.subheader("جدول مقایسه")

if "sort_by" not in st.session_state:
    st.session_state.sort_by = "P/E Forward"
    st.session_state.sort_asc = True

def safe_sort(dataframe, column, ascending=True):
    temp = dataframe.copy()
    values = []
    for v in temp[column]:
        if v is None:
            values.append(1e18 if ascending else -1e18)
        else:
            try:
                values.append(float(v))
            except Exception:
                values.append(1e18 if ascending else -1e18)
    temp["_sort_key"] = values
    temp = temp.sort_values("_sort_key", ascending=ascending).reset_index(drop=True)
    return temp.drop(columns=["_sort_key"])

# ---- کنترل مرتب‌سازی (تمیز و یک‌خطی) ----
sort_options = {
    "P/E Forward": "P/E Forward",
    "برآورد سود": "برآورد سود",
    "ارزش بازار": "ارزش بازار",
    "میانگین حاشیه": "میانگین حاشیه",
    "برآورد فروش": "برآورد فروش",
    "P/S Forward": "P/S Forward",
    "P/D Forward": "P/D Forward",
}

s1, s2, s3 = st.columns([3, 2, 5])
with s1:
    chosen = st.selectbox(
        "مرتب‌سازی بر اساس:",
        options=list(sort_options.keys()),
        index=list(sort_options.keys()).index(
            st.session_state.sort_by if st.session_state.sort_by in sort_options else "P/E Forward"
        )
    )
    st.session_state.sort_by = sort_options[chosen]
with s2:
    direction = st.radio(
        "جهت:",
        options=["صعودی ▲", "نزولی ▼"],
        index=0 if st.session_state.sort_asc else 1,
        horizontal=True
    )
    st.session_state.sort_asc = direction.startswith("صعودی")

df_view = safe_sort(df, st.session_state.sort_by, st.session_state.sort_asc)

# ---- جدول HTML مرتب ----
html = """
<table style="width:100%; border-collapse:collapse; direction:rtl; text-align:center; font-size:13px;">
<thead>
<tr style="background:#1e3a5f; color:white;">
    <th style="padding:8px; border:1px solid #ccc;">نماد</th>
    <th style="padding:8px; border:1px solid #ccc;">نام</th>
    <th style="padding:8px; border:1px solid #ccc;">صنعت</th>
    <th style="padding:8px; border:1px solid #ccc;">ارزش بازار</th>
    <th style="padding:8px; border:1px solid #ccc;">برآورد فروش</th>
    <th style="padding:8px; border:1px solid #ccc;">میانگین حاشیه</th>
    <th style="padding:8px; border:1px solid #ccc;">برآورد سود</th>
    <th style="padding:8px; border:1px solid #ccc;">سود تقسیمی</th>
    <th style="padding:8px; border:1px solid #ccc;">P/E Forward</th>
    <th style="padding:8px; border:1px solid #ccc;">P/D Forward</th>
    <th style="padding:8px; border:1px solid #ccc;">P/S Forward</th>
    <th style="padding:8px; border:1px solid #ccc;">حاشیه آخرین دوره</th>
    <th style="padding:8px; border:1px solid #ccc;">نسبت مطالبات</th>
</tr>
</thead>
<tbody>
"""

for i, (_, row) in enumerate(df_view.iterrows()):
    bg = "#f8fafc" if i % 2 == 0 else "#ffffff"
    pe_cell = pe_html(row["P/E Forward"])
    html += f"""
<tr style="background:{bg};">
    <td style="padding:7px; border:1px solid #ddd; font-weight:bold;">{row['نماد']}</td>
    <td style="padding:7px; border:1px solid #ddd;">{row['نام']}</td>
    <td style="padding:7px; border:1px solid #ddd;">{row['صنعت']}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['ارزش بازار'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['برآورد فروش'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_pct(row['میانگین حاشیه'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['برآورد سود'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['برآورد سود تقسیمی'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{pe_cell}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_ratio(row['P/D Forward'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_ratio(row['P/S Forward'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_pct(row['حاشیه آخرین دوره'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_pct(row['نسبت مطالبات'])}</td>
</tr>
"""

html += "</tbody></table>"
st.markdown(html, unsafe_allow_html=True)
st.caption("P/E: 🟢 مناسب < ۵  |  🟡 متوسط ۵ تا ۷  |  🔴 بالا > ۷")
# ============================================================
# خروجی اکسل با فرمت‌بندی
# ============================================================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

def build_excel(df_export):
    wb = Workbook()
    ws = wb.active
    ws.title = "مقایسه شرکت‌ها"
    ws.sheet_view.rightToLeft = True

    headers = [
        "نماد", "نام", "صنعت",
        "ارزش بازار", "برآورد فروش", "میانگین حاشیه (%)",
        "برآورد سود", "سود تقسیمی",
        "P/E Forward", "P/D Forward", "P/S Forward",
        "حاشیه آخرین دوره (%)", "نسبت مطالبات (%)"
    ]

    # استایل‌ها
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    # هدر
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    ws.row_dimensions[1].height = 28

    # داده‌ها
    for r_idx, (_, row) in enumerate(df_export.iterrows(), 2):
        values = [
            row["نماد"],
            row["نام"],
            row["صنعت"],
            row["ارزش بازار"],
            row["برآورد فروش"],
            row["میانگین حاشیه"],          # بعداً ×100 می‌شود
            row["برآورد سود"],
            row["برآورد سود تقسیمی"],
            row["P/E Forward"],
            row["P/D Forward"],
            row["P/S Forward"],
            row["حاشیه آخرین دوره"],
            row["نسبت مطالبات"],
        ]

        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = cell_align
            cell.border = thin
            if r_idx % 2 == 0:
                cell.fill = alt_fill

            # ستون‌های متنی
            if c_idx in (1, 2, 3):
                cell.value = val if val is not None else "—"
                continue

            # اعداد بزرگ با جداکننده هزارگان (بدون اعشار)
            if c_idx in (4, 5, 7, 8):  # ارزش بازار، برآورد فروش، برآورد سود، سود تقسیمی
                if val is None:
                    cell.value = "—"
                else:
                    cell.value = float(val)
                    cell.number_format = '#,##0'

            # درصدها (دو رقم اعشار)
            elif c_idx in (6, 12, 13):  # میانگین حاشیه، حاشیه آخرین، نسبت مطالبات
                if val is None:
                    cell.value = "—"
                else:
                    cell.value = float(val)   # مقدار بین 0 و 1
                    cell.number_format = '0.00%'

            # نسبت‌ها با دو رقم اعشار
            elif c_idx in (9, 10, 11):  # P/E, P/D, P/S
                if val is None:
                    cell.value = "—"
                else:
                    cell.value = float(val)
                    cell.number_format = '0.00'
                    # رنگ شرطی فقط برای P/E
                    if c_idx == 9:
                        pe = float(val)
                        if pe < 5:
                            cell.fill = green_fill
                        elif pe <= 7:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill

    # عرض ستون‌ها
    widths = [10, 22, 18, 14, 14, 14, 14, 12, 12, 12, 12, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # فریز هدر
    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

try:
    excel_buffer = build_excel(df_view)
    st.download_button(
        label="📥 دانلود جدول مقایسه (Excel فرمت‌بندی‌شده)",
        data=excel_buffer,
        file_name="moghayese_sherkatha.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
except Exception as e:
    st.warning(f"خطا در ساخت اکسل: {e}")
    # fallback CSV
    csv_df = df_view.copy()
    for col in ["میانگین حاشیه", "حاشیه آخرین دوره", "نسبت مطالبات"]:
        csv_df[col] = csv_df[col].apply(lambda v: round(v * 100, 2) if v is not None else None)
    for col in ["P/E Forward", "P/D Forward", "P/S Forward"]:
        csv_df[col] = csv_df[col].apply(lambda v: round(v, 2) if v is not None else None)
    csv_buffer = BytesIO()
    csv_df.to_csv(csv_buffer, index=False, encoding="utf-8-sig")
    csv_buffer.seek(0)
    st.download_button(
        label="دانلود CSV",
        data=csv_buffer,
        file_name="moghayese_sherkatha.csv",
        mime="text/csv"
    )