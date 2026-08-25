# pages/APi1.py
# -*- coding: utf-8 -*-

import streamlit as st
import sys
import os
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import requests
import json
import time
from datetime import datetime, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# =====================================================
# =============== اعمال استایل سراسری =================
# =====================================================

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from utils.styles import apply_styles
apply_styles()

# =====================================================
# =============== تنظیمات پیش‌فرض =====================
# =====================================================

DEFAULT_CONFIG = {
    "QUALITY_SCORES": {
        "PE": {
            "weight": 30,
            "excellent": 7,
            "good": 10,
        },
        "EPS": {
            "weight": 20,
            "excellent": 500,
            "good": 300,
            "medium": 100,
        },
        "VOLUME": {
            "weight": 20,
            "excellent": 5000000,
            "good": 1000000,
        },
        "PRICE_CHANGE": {
            "weight": 20,
            "excellent": 5,
            "good": 0,
        },
        "MARKET_VALUE": {
            "weight": 15,
            "excellent_min": 0.1,
            "excellent_max": 10,
            "good_max": 0.1,
        }
    },
    "MONEY_FLOW": {
        "strong_inflow": 80,
        "inflow": 60,
        "outflow": 20,
        "strong_outflow": 10,
    },
    "PE_COLORS": {
        "excellent": 5,
        "good": 7,
    },
    "DEFAULT_FILTERS": {
        "min_quality": 30,
        "max_pe": 15,
        "min_volume": 100000,
    },
    "DISPLAY": {
        "top_companies": 10,
        "table_rows": 100,
        "top_money_flow": 20,
    }
}

# =====================================================
# =============== مقداردهی اولیه تنظیمات ==============
# =====================================================

if 'config' not in st.session_state:
    st.session_state.config = DEFAULT_CONFIG.copy()

if 'config_initialized' not in st.session_state:
    st.session_state.config_initialized = True

# =====================================================
# =============== استایل‌های اختصاصی ==================
# =====================================================

st.markdown("""
<style>
    .main .block-container {
        padding-top: 1.2rem !important;
        padding-right: 2rem !important;
        padding-left: 2rem !important;
        max-width: 100% !important;
    }
    
    .config-section {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 1.5rem;
        margin-bottom: 1.5rem;
    }
    
    .config-section h4 {
        color: #0f172a;
        font-weight: 600;
        margin-bottom: 1rem;
        font-family: 'Vazirmatn', Tahoma, sans-serif;
    }
    
    .config-section .stNumberInput {
        direction: ltr;
    }
    
    .search-section {
        background: #f8fafc;
        border: 1px solid #e2e8f0;
        border-radius: 12px;
        padding: 1.2rem;
        margin-bottom: 1.5rem;
    }
    
    .search-section .stTextInput {
        direction: rtl;
    }
    
    .stTabs {
        margin-bottom: 1.5rem !important;
    }
    
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
        background-color: #f1f5f9;
        padding: 8px;
        border-radius: 12px;
        flex-wrap: wrap;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        padding: 10px 20px;
        font-weight: 500;
        font-family: 'Vazirmatn', Tahoma, sans-serif !important;
        font-size: 0.95rem !important;
        color: #475569 !important;
        background: transparent !important;
        transition: all 0.3s ease !important;
        border: none !important;
    }
    
    .stTabs [data-baseweb="tab"]:hover {
        background: #e2e8f0 !important;
        color: #0f172a !important;
    }
    
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%) !important;
        color: #ffffff !important;
        box-shadow: 0 2px 8px rgba(15, 23, 42, 0.15) !important;
    }
    
    .stTabs [data-baseweb="tab-panel"] {
        padding-top: 1.5rem !important;
        background: #ffffff !important;
        border-radius: 0 0 12px 12px !important;
    }
    
    .custom-table {
        width: 100%;
        border-collapse: collapse;
        direction: rtl;
        text-align: center;
        font-size: 13px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.05);
        border-radius: 8px;
        overflow: hidden;
        font-family: 'Vazirmatn', Tahoma, sans-serif !important;
        margin: 0.5rem 0;
    }
    
    .custom-table thead tr {
        background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
        color: white;
        font-weight: 600;
    }
    
    .custom-table th {
        padding: 12px 8px;
        border: 1px solid #334155;
        font-size: 13px;
        font-weight: 600;
        white-space: nowrap;
    }
    
    .custom-table td {
        padding: 10px 8px;
        border: 1px solid #e2e8f0;
        font-size: 13px;
    }
    
    .custom-table tbody tr:nth-child(even) {
        background-color: #f8fafc;
    }
    
    .custom-table tbody tr:nth-child(odd) {
        background-color: #ffffff;
    }
    
    .custom-table tbody tr:hover {
        background-color: #f1f5f9 !important;
        transition: background-color 0.2s;
    }
    
    .footer-text {
        text-align: center;
        font-size: 0.95rem !important;
        color: #334155 !important;
        font-family: 'Vazirmatn', Tahoma, sans-serif !important;
        padding: 1rem 0;
        border-top: 2px solid #f1f5f9;
        margin-top: 2rem;
    }
    .footer-text .highlight {
        color: #0f172a;
        font-weight: 600;
    }
    .footer-text .divider {
        color: #cbd5e1;
        margin: 0 0.5rem;
    }
    
    @media (max-width: 768px) {
        .custom-table {
            font-size: 11px;
        }
        .custom-table th, .custom-table td {
            padding: 6px 4px;
        }
        .main .block-container {
            padding-right: 0.5rem !important;
            padding-left: 0.5rem !important;
        }
        .stTabs [data-baseweb="tab"] {
            padding: 6px 12px !important;
            font-size: 0.85rem !important;
        }
    }
</style>
""", unsafe_allow_html=True)

st.title("📊 تحلیل پیشرفته بازار - ورود و خروج پول")
st.markdown("---")

# =====================================================
# =============== توابع دریافت داده ====================
# =====================================================

@st.cache(ttl=3600, allow_output_mutation=True)
def fetch_data_from_api():
    url = "https://api.brsapi.ir/Tsetmc/AllSymbols.php?key=B9a9DpJKnmEAbgXDVkzh3kL3f7EzVSsd&type=1"
    
    try:
        session = requests.Session()
        session.trust_env = False
        session.verify = False
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36'
        }
        
        response = session.get(url, headers=headers, timeout=60)
        
        if response.status_code == 200:
            return response.json()
        else:
            return None
    except Exception as e:
        st.error(f"❌ خطا در دریافت داده: {e}")
        return None

def load_cached_data():
    try:
        with open('analysis_cache.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        return None

def save_to_cache(data):
    try:
        with open('analysis_cache.json', 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
        return True
    except:
        return False

# =====================================================
# =============== توابع کمکی ==========================
# =====================================================

PERSIAN_DIGITS = {'0': '۰', '1': '۱', '2': '۲', '3': '۳', '4': '۴', 
                  '5': '۵', '6': '۶', '7': '۷', '8': '۸', '9': '۹'}

def to_persian_number(num):
    if num is None or num == "" or (isinstance(num, float) and num != num):
        return "—"
    try:
        if isinstance(num, float):
            if num == int(num):
                num_str = f"{int(num):,}"
            else:
                num_str = f"{num:,.2f}"
        else:
            num_str = f"{int(num):,}"
        
        for en, fa in PERSIAN_DIGITS.items():
            num_str = num_str.replace(en, fa)
        return num_str
    except:
        return str(num)

def fmt(v):
    if v is None:
        return "—"
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return "—"

def fmt_market_value_persian(v):
    if v is None:
        return "—"
    try:
        value_billion = float(v) / 10_000_000_000
        num_str = f"{value_billion:,.1f}"
        for en, fa in PERSIAN_DIGITS.items():
            num_str = num_str.replace(en, fa)
        return num_str
    except Exception:
        return "—"

def pe_color(pe, config):
    if pe is None:
        return "#6b7280"
    excellent = config["PE_COLORS"]["excellent"]
    good = config["PE_COLORS"]["good"]
    if pe < excellent:
        return "#16a34a"
    if pe <= good:
        return "#ca8a04"
    return "#dc2626"

def pe_label(pe):
    if pe is None:
        return "—"
    if pe < 5:
        return f"{pe:.2f} (مناسب)"
    if pe <= 7:
        return f"{pe:.2f} (متوسط)"
    return f"{pe:.2f} (بالا)"

def get_quality_label(score):
    if score >= 80:
        return "⭐ عالی"
    elif score >= 60:
        return "🌟 خوب"
    elif score >= 40:
        return "👍 متوسط"
    else:
        return "⚠️ ضعیف"

# =====================================================
# =============== تشخیص ورود/خروج پول =================
# =====================================================

def detect_money_flow_advanced(row, config):
    """
    تشخیص ورود/خروج پول بر اساس حجم و قیمت
    """
    today_volume = row.get('tvol', 0)
    price_change = row.get('plp', 0)
    
    # محاسبه میانگین حجم (تخمینی)
    monthly_avg = row.get('tvol_avg_1m', None)
    
    if monthly_avg is None or monthly_avg == 0:
        # اگر میانگین نداشتیم، از روش رتبه‌بندی استفاده می‌کنیم
        return get_money_flow_status(row.get('money_flow_rank', 0), config)
    
    # محاسبه نسبت حجم به میانگین
    volume_ratio = today_volume / monthly_avg if monthly_avg > 0 else 0
    
    # تشخیص بر اساس ترکیب حجم و قیمت
    if volume_ratio >= 2:
        if price_change > 0:
            return '🟢 ورود پول قوی'
        elif price_change < 0:
            return '🔴 خروج پول قوی'
        else:
            return '🟡 حجم بالا، قیمت ثابت'
    
    elif volume_ratio >= 1.5:
        if price_change > 0:
            return '🟡 ورود پول'
        elif price_change < 0:
            return '⚪ خروج پول'
        else:
            return '🔵 حجم متوسط، قیمت ثابت'
    
    else:
        if price_change > 0:
            return '🔵 رشد بدون حجم'
        elif price_change < 0:
            return '🔵 ریزش بدون حجم'
        else:
            return '🔵 عادی'

def get_money_flow_status(rank, config):
    """تعیین وضعیت جریان پول بر اساس تنظیمات"""
    mf = config["MONEY_FLOW"]
    if rank > mf["strong_inflow"]:
        return '🟢 ورود پول قوی'
    elif rank > mf["inflow"]:
        return '🟡 ورود پول'
    elif rank < mf["strong_outflow"]:
        return '⛔ خروج پول قوی'
    elif rank < mf["outflow"]:
        return '⚪ خروج پول'
    else:
        return '🔵 متوسط'

# =====================================================
# =============== توابع محاسبه با تنظیمات =============
# =====================================================

def calculate_quality_score_with_config(row, config):
    """محاسبه امتیاز کیفیت بر اساس تنظیمات"""
    score = 0
    qs = config["QUALITY_SCORES"]
    
    # P/E
    pe = row.get('pe', 0)
    if 0 < pe < qs["PE"]["excellent"]:
        score += qs["PE"]["weight"]
    elif qs["PE"]["excellent"] <= pe <= qs["PE"]["good"]:
        score += qs["PE"]["weight"] / 2
    
    # EPS
    eps = row.get('eps', 0)
    if eps > qs["EPS"]["excellent"]:
        score += qs["EPS"]["weight"]
    elif eps > qs["EPS"]["good"]:
        score += qs["EPS"]["weight"] * 0.75
    elif eps > qs["EPS"]["medium"]:
        score += qs["EPS"]["weight"] * 0.5
    
    # حجم معاملات
    volume = row.get('tvol', 0)
    if volume > qs["VOLUME"]["excellent"]:
        score += qs["VOLUME"]["weight"]
    elif volume > qs["VOLUME"]["good"]:
        score += qs["VOLUME"]["weight"] / 2
    
    # رشد قیمت
    if row.get('plp', 0) > qs["PRICE_CHANGE"]["excellent"]:
        score += qs["PRICE_CHANGE"]["weight"]
    elif row.get('plp', 0) > qs["PRICE_CHANGE"]["good"]:
        score += qs["PRICE_CHANGE"]["weight"] * 0.75
    
    # ارزش بازار
    mv = row.get('mv', 0) / 1_000_000_000_000
    if qs["MARKET_VALUE"]["excellent_min"] < mv < qs["MARKET_VALUE"]["excellent_max"]:
        score += qs["MARKET_VALUE"]["weight"]
    elif mv <= qs["MARKET_VALUE"]["good_max"]:
        score += qs["MARKET_VALUE"]["weight"] * 0.67
    
    return min(score, 100)

# =====================================================
# =============== دکمه‌های دریافت داده ================
# =====================================================

col1, col2, col3 = st.columns(3)

with col1:
    if st.button("🔄 دریافت داده از API"):
        with st.spinner("در حال دریافت داده..."):
            data = fetch_data_from_api()
            if data:
                save_to_cache(data)
                st.session_state['market_data'] = data
                st.success(f"✅ {len(data)} نماد دریافت شد!")
                st.experimental_rerun()
            else:
                st.error("❌ دریافت داده ناموفق")
                cached = load_cached_data()
                if cached:
                    st.session_state['market_data'] = cached
                    st.warning("⚠️ استفاده از داده‌های کش شده")
                    st.experimental_rerun()

with col2:
    if st.button("📂 بارگذاری از کش"):
        cached = load_cached_data()
        if cached:
            st.session_state['market_data'] = cached
            st.success(f"✅ {len(cached)} نماد از کش بارگذاری شد!")
            st.experimental_rerun()
        else:
            st.error("❌ فایل کش پیدا نشد")

with col3:
    if st.button("🧪 داده‌های نمونه"):
        import random
        symbols = ['فولاد', 'خودرو', 'شپنا', 'فملی', 'کگل', 'وبملت', 'شستا', 
                   'پارس', 'سایپا', 'ریس', 'خگستر', 'وتجارت', 'حفاری', 'پخش']
        industries = ['فولاد', 'خودرو', 'پالایشی', 'معدنی', 'بانکی', 'سیمانی', 'دارویی']
        data = []
        for i in range(50):
            symbol = random.choice(symbols) + str(random.randint(1, 99))
            data.append({
                'l18': symbol,
                'l30': f'شرکت {symbol}',
                'cs': random.choice(industries),
                'mv': random.randint(100_000_000_000, 50_000_000_000_000),
                'pc': random.uniform(100, 5000),
                'plp': random.uniform(-10, 15),
                'pe': random.uniform(2, 25),
                'eps': random.randint(10, 1000),
                'tvol': random.randint(100_000, 10_000_000)
            })
        st.session_state['market_data'] = data
        st.success(f"✅ {len(data)} نماد نمونه تولید شد!")
        st.experimental_rerun()

st.markdown("---")

# =====================================================
# =============== بررسی داده ==========================
# =====================================================

if 'market_data' not in st.session_state:
    st.info("👈 برای شروع، روی دکمه 'دریافت داده از API' کلیک کنید.")
    st.stop()

df = pd.DataFrame(st.session_state['market_data'])

# =====================================================
# =============== ادامه کد محاسبات ====================
# =====================================================

# دریافت تنظیمات جاری
config = st.session_state.config

# محاسبه شاخص‌ها با تنظیمات جدید
df['money_flow'] = df['tvol'] * df['pc']
df['money_flow_rank'] = df['money_flow'].rank(pct=True) * 100
df['mv_billion'] = df['mv'] / 10_000_000_000

# محاسبه کیفیت با تنظیمات جدید
df['quality_score'] = df.apply(lambda row: calculate_quality_score_with_config(row, config), axis=1)
df['quality_label'] = df['quality_score'].apply(get_quality_label)

# تشخیص ورود/خروج پول با روش جدید (ترکیبی)
df['money_flow_status'] = df.apply(lambda row: detect_money_flow_advanced(row, config), axis=1)
df['overall_rank'] = df['quality_score'].rank(method='min', ascending=False)

# =====================================================
# =============== بخش جستجو و فیلتر ===================
# =====================================================

st.markdown('<div class="search-section">', unsafe_allow_html=True)
st.subheader("🔍 جستجو و فیلتر نمادها")

# =============================================
# انتخاب روش جستجو
# =============================================
search_col1, search_col2, search_col3 = st.columns([2, 1, 1])

with search_col1:
    search_mode = st.radio(
        "روش جستجو:",
        options=['🔍 بر اساس نماد', '🏭 بر اساس صنعت', '📊 همه نمادها'],
        horizontal=True
    )

with search_col2:
    search_input = ""
    if search_mode == '🔍 بر اساس نماد':
        search_input = st.text_input("جستجوی نماد:", placeholder="مثال: شپنا, فولاد")
    elif search_mode == '🏭 بر اساس صنعت':
        industries = df['cs'].unique().tolist() if 'cs' in df.columns else []
        selected_industry = st.selectbox("انتخاب صنعت:", ['همه'] + industries)
        search_input = selected_industry

with search_col3:
    # دکمه پاک کردن فیلتر
    if st.button("🗑️ پاک کردن فیلتر"):
        st.experimental_rerun()

st.markdown('</div>', unsafe_allow_html=True)

# =============================================
# اعمال فیلتر
# =============================================
filtered_df = df.copy()

if search_mode == '🔍 بر اساس نماد' and search_input:
    filtered_df = df[df['l18'].str.contains(search_input.upper(), na=False)]
elif search_mode == '🏭 بر اساس صنعت' and search_input and search_input != 'همه':
    if 'cs' in df.columns:
        filtered_df = df[df['cs'] == search_input]

# نمایش تعداد نتایج
st.caption(f"📊 تعداد نمادهای یافت‌شده: **{len(filtered_df)}**")
st.markdown("---")

# =====================================================
# =============== تعریف تب‌ها =========================
# =====================================================

tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs([
    "📋 جدول تحلیل",
    "⭐ بهترین‌ها",
    "💰 جریان پول",
    "📈 تحلیل حجم",
    "📥 خروجی اکسل",
    "⚙️ تنظیمات شاخص‌ها"
])

# =====================================================
# =============== تب 6: تنظیمات شاخص‌ها ===============
# =====================================================

with tab6:
    st.subheader("⚙️ تنظیمات شاخص‌های تحلیل")
    st.markdown("""
    در این بخش می‌توانید وزن و آستانه‌های هر شاخص را تغییر دهید.  
    تغییرات به صورت **همزمان** روی تمام تحلیل‌ها اعمال می‌شود.
    """)
    
    # دکمه بازنشانی به تنظیمات پیش‌فرض
    col1, col2 = st.columns([3, 1])
    with col2:
        if st.button("🔄 بازنشانی به پیش‌فرض"):
            st.session_state.config = DEFAULT_CONFIG.copy()
            st.success("✅ تنظیمات به حالت پیش‌فرض بازنشانی شد!")
            st.experimental_rerun()
    
    st.markdown("---")
    
    # ==========================================
    # بخش 1: وزن شاخص‌ها
    # ==========================================
    st.markdown("### 📊 وزن شاخص‌های کیفیت")
    st.markdown("مجموع وزن‌ها باید ۱۰۰ باشد.")
    
    col1, col2, col3, col4, col5 = st.columns(5)
    with col1:
        pe_weight = st.number_input(
            "وزن P/E",
            min_value=0,
            max_value=100,
            value=int(st.session_state.config["QUALITY_SCORES"]["PE"]["weight"]),
            step=5,
            key="pe_weight"
        )
    with col2:
        eps_weight = st.number_input(
            "وزن EPS",
            min_value=0,
            max_value=100,
            value=int(st.session_state.config["QUALITY_SCORES"]["EPS"]["weight"]),
            step=5,
            key="eps_weight"
        )
    with col3:
        vol_weight = st.number_input(
            "وزن حجم",
            min_value=0,
            max_value=100,
            value=int(st.session_state.config["QUALITY_SCORES"]["VOLUME"]["weight"]),
            step=5,
            key="vol_weight"
        )
    with col4:
        price_weight = st.number_input(
            "وزن رشد قیمت",
            min_value=0,
            max_value=100,
            value=int(st.session_state.config["QUALITY_SCORES"]["PRICE_CHANGE"]["weight"]),
            step=5,
            key="price_weight"
        )
    with col5:
        mv_weight = st.number_input(
            "وزن ارزش بازار",
            min_value=0,
            max_value=100,
            value=int(st.session_state.config["QUALITY_SCORES"]["MARKET_VALUE"]["weight"]),
            step=5,
            key="mv_weight"
        )
    
    total_weight = pe_weight + eps_weight + vol_weight + price_weight + mv_weight
    if total_weight != 100:
        st.warning(f"⚠️ مجموع وزن‌ها باید ۱۰۰ باشد. مقدار فعلی: {total_weight}")
    else:
        st.session_state.config["QUALITY_SCORES"]["PE"]["weight"] = pe_weight
        st.session_state.config["QUALITY_SCORES"]["EPS"]["weight"] = eps_weight
        st.session_state.config["QUALITY_SCORES"]["VOLUME"]["weight"] = vol_weight
        st.session_state.config["QUALITY_SCORES"]["PRICE_CHANGE"]["weight"] = price_weight
        st.session_state.config["QUALITY_SCORES"]["MARKET_VALUE"]["weight"] = mv_weight
        st.success(f"✅ وزن‌ها با موفقیت ذخیره شدند (مجموع: {total_weight})")
    
    st.markdown("---")
    
    # ==========================================
    # بخش 2: آستانه‌های شاخص‌ها
    # ==========================================
    st.markdown("### 🎯 آستانه‌های شاخص‌ها")
    
    # P/E - همه int
    with st.expander("📊 P/E (نسبت قیمت به سود)", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            pe_excellent = st.number_input(
                "P/E عالی (کمتر از)",
                min_value=1,
                max_value=20,
                value=int(st.session_state.config["QUALITY_SCORES"]["PE"]["excellent"]),
                step=1,
                key="pe_excellent"
            )
        with col2:
            pe_good = st.number_input(
                "P/E خوب (کمتر از)",
                min_value=int(pe_excellent),
                max_value=30,
                value=int(st.session_state.config["QUALITY_SCORES"]["PE"]["good"]),
                step=1,
                key="pe_good"
            )
        st.session_state.config["QUALITY_SCORES"]["PE"]["excellent"] = int(pe_excellent)
        st.session_state.config["QUALITY_SCORES"]["PE"]["good"] = int(pe_good)
        st.info(f"🟢 P/E < {int(pe_excellent)}: عالی | 🟡 {int(pe_excellent)} ≤ P/E ≤ {int(pe_good)}: خوب | 🔴 P/E > {int(pe_good)}: ضعیف")
    
    # EPS - همه int
    with st.expander("💰 EPS (سود هر سهم)", expanded=False):
        col1, col2, col3 = st.columns(3)
        with col1:
            eps_excellent = st.number_input(
                "EPS عالی (بیشتر از)",
                min_value=0,
                max_value=2000,
                value=int(st.session_state.config["QUALITY_SCORES"]["EPS"]["excellent"]),
                step=50,
                key="eps_excellent"
            )
        with col2:
            eps_good = st.number_input(
                "EPS خوب (بیشتر از)",
                min_value=0,
                max_value=int(eps_excellent),
                value=int(st.session_state.config["QUALITY_SCORES"]["EPS"]["good"]),
                step=50,
                key="eps_good"
            )
        with col3:
            eps_medium = st.number_input(
                "EPS متوسط (بیشتر از)",
                min_value=0,
                max_value=int(eps_good),
                value=int(st.session_state.config["QUALITY_SCORES"]["EPS"]["medium"]),
                step=50,
                key="eps_medium"
            )
        st.session_state.config["QUALITY_SCORES"]["EPS"]["excellent"] = int(eps_excellent)
        st.session_state.config["QUALITY_SCORES"]["EPS"]["good"] = int(eps_good)
        st.session_state.config["QUALITY_SCORES"]["EPS"]["medium"] = int(eps_medium)
        st.info(f"🟢 EPS > {int(eps_excellent)}: عالی | 🟡 {int(eps_good)} < EPS ≤ {int(eps_excellent)}: خوب | 🟠 {int(eps_medium)} < EPS ≤ {int(eps_good)}: متوسط | 🔴 EPS ≤ {int(eps_medium)}: ضعیف")
    
    # حجم معاملات - همه int
    with st.expander("📊 حجم معاملات", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            vol_excellent = st.number_input(
                "حجم عالی (بیشتر از)",
                min_value=0,
                max_value=100000000,
                value=int(st.session_state.config["QUALITY_SCORES"]["VOLUME"]["excellent"]),
                step=500000,
                key="vol_excellent"
            )
        with col2:
            vol_good = st.number_input(
                "حجم خوب (بیشتر از)",
                min_value=0,
                max_value=int(vol_excellent),
                value=int(st.session_state.config["QUALITY_SCORES"]["VOLUME"]["good"]),
                step=500000,
                key="vol_good"
            )
        st.session_state.config["QUALITY_SCORES"]["VOLUME"]["excellent"] = int(vol_excellent)
        st.session_state.config["QUALITY_SCORES"]["VOLUME"]["good"] = int(vol_good)
        st.info(f"🟢 حجم > {int(vol_excellent):,}: عالی | 🟡 {int(vol_good):,} < حجم ≤ {int(vol_excellent):,}: خوب | 🔴 حجم ≤ {int(vol_good):,}: ضعیف")
    
    # رشد قیمت - همه int
    with st.expander("📈 رشد قیمت", expanded=False):
        col1, col2 = st.columns(2)
        with col1:
            price_excellent = st.number_input(
                "رشد عالی (بیشتر از %)",
                min_value=-10,
                max_value=30,
                value=int(st.session_state.config["QUALITY_SCORES"]["PRICE_CHANGE"]["excellent"]),
                step=1,
                key="price_excellent"
            )
        with col2:
            price_good = st.number_input(
                "رشد خوب (بیشتر از %)",
                min_value=-10,
                max_value=int(price_excellent),
                value=int(st.session_state.config["QUALITY_SCORES"]["PRICE_CHANGE"]["good"]),
                step=1,
                key="price_good"
            )
        st.session_state.config["QUALITY_SCORES"]["PRICE_CHANGE"]["excellent"] = int(price_excellent)
        st.session_state.config["QUALITY_SCORES"]["PRICE_CHANGE"]["good"] = int(price_good)
        st.info(f"🟢 رشد > {int(price_excellent)}%: عالی | 🟡 {int(price_good)}% < رشد ≤ {int(price_excellent)}%: خوب | 🔴 رشد ≤ {int(price_good)}%: ضعیف")
    
    # ارزش بازار - همه float
    with st.expander("🏢 ارزش بازار", expanded=False):
        col1, col2, col3 = st.columns(3)
        with col1:
            mv_min = st.number_input(
                "حداقل عالی (تریلیون)",
                min_value=0.0,
                max_value=50.0,
                value=float(st.session_state.config["QUALITY_SCORES"]["MARKET_VALUE"]["excellent_min"]),
                step=0.1,
                key="mv_min"
            )
        with col2:
            mv_max = st.number_input(
                "حداکثر عالی (تریلیون)",
                min_value=float(mv_min),
                max_value=100.0,
                value=float(st.session_state.config["QUALITY_SCORES"]["MARKET_VALUE"]["excellent_max"]),
                step=0.1,
                key="mv_max"
            )
        with col3:
            mv_good = st.number_input(
                "حداکثر خوب (تریلیون)",
                min_value=0.0,
                max_value=float(mv_min),
                value=float(st.session_state.config["QUALITY_SCORES"]["MARKET_VALUE"]["good_max"]),
                step=0.1,
                key="mv_good"
            )
        st.session_state.config["QUALITY_SCORES"]["MARKET_VALUE"]["excellent_min"] = float(mv_min)
        st.session_state.config["QUALITY_SCORES"]["MARKET_VALUE"]["excellent_max"] = float(mv_max)
        st.session_state.config["QUALITY_SCORES"]["MARKET_VALUE"]["good_max"] = float(mv_good)
        st.info(f"🟢 {float(mv_min):.1f} < MV < {float(mv_max):.1f} تریلیون: عالی | 🟡 MV ≤ {float(mv_good):.1f} تریلیون: خوب | 🔴 MV ≥ {float(mv_max):.1f} تریلیون: ضعیف")
    
    st.markdown("---")
    
    # ==========================================
    # بخش 3: تنظیمات جریان پول - همه int
    # ==========================================
    st.markdown("### 💰 تنظیمات جریان پول")
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        strong_inflow = st.number_input(
            "ورود پول قوی (%)",
            min_value=0,
            max_value=100,
            value=int(st.session_state.config["MONEY_FLOW"]["strong_inflow"]),
            step=1,
            key="strong_inflow"
        )
    with col2:
        inflow = st.number_input(
            "ورود پول (%)",
            min_value=0,
            max_value=int(strong_inflow),
            value=int(st.session_state.config["MONEY_FLOW"]["inflow"]),
            step=1,
            key="inflow"
        )
    with col3:
        outflow = st.number_input(
            "خروج پول (%)",
            min_value=0,
            max_value=int(inflow),
            value=int(st.session_state.config["MONEY_FLOW"]["outflow"]),
            step=1,
            key="outflow"
        )
    with col4:
        strong_outflow = st.number_input(
            "خروج پول قوی (%)",
            min_value=0,
            max_value=int(outflow),
            value=int(st.session_state.config["MONEY_FLOW"]["strong_outflow"]),
            step=1,
            key="strong_outflow"
        )
    
    st.session_state.config["MONEY_FLOW"]["strong_inflow"] = int(strong_inflow)
    st.session_state.config["MONEY_FLOW"]["inflow"] = int(inflow)
    st.session_state.config["MONEY_FLOW"]["outflow"] = int(outflow)
    st.session_state.config["MONEY_FLOW"]["strong_outflow"] = int(strong_outflow)
    
    st.info(f"""
    🟢 ورود پول قوی: > {int(strong_inflow)}%  
    🟡 ورود پول: {int(inflow)}% - {int(strong_inflow)}%  
    🔵 متوسط: {int(outflow)}% - {int(inflow)}%  
    ⚪ خروج پول: {int(strong_outflow)}% - {int(outflow)}%  
    ⛔ خروج پول قوی: < {int(strong_outflow)}%
    """)
    
    st.markdown("---")
    
    # ==========================================
    # بخش 4: تنظیمات P/E رنگ‌بندی - همه int
    # ==========================================
    st.markdown("### 🎨 رنگ‌بندی P/E")
    
    col1, col2 = st.columns(2)
    with col1:
        pe_green = st.number_input(
            "حداکثر P/E سبز (مناسب)",
            min_value=1,
            max_value=20,
            value=int(st.session_state.config["PE_COLORS"]["excellent"]),
            step=1,
            key="pe_green"
        )
    with col2:
        pe_yellow = st.number_input(
            "حداکثر P/E زرد (متوسط)",
            min_value=int(pe_green),
            max_value=30,
            value=int(st.session_state.config["PE_COLORS"]["good"]),
            step=1,
            key="pe_yellow"
        )
    
    st.session_state.config["PE_COLORS"]["excellent"] = int(pe_green)
    st.session_state.config["PE_COLORS"]["good"] = int(pe_yellow)
    
    st.markdown(f"""
    🟢 **سبز (مناسب)**: P/E < {int(pe_green)}  
    🟡 **زرد (متوسط)**: {int(pe_green)} ≤ P/E ≤ {int(pe_yellow)}  
    🔴 **قرمز (بالا)**: P/E > {int(pe_yellow)}
    """)
    
    st.markdown("---")
    
    # ==========================================
    # بخش 5: فیلترهای پیش‌فرض - همه int
    # ==========================================
    st.markdown("### 🔍 فیلترهای پیش‌فرض")
    
    col1, col2, col3 = st.columns(3)
    with col1:
        default_min_quality = st.number_input(
            "حداقل کیفیت پیش‌فرض",
            min_value=0,
            max_value=100,
            value=int(st.session_state.config["DEFAULT_FILTERS"]["min_quality"]),
            step=5,
            key="default_min_quality"
        )
    with col2:
        default_max_pe = st.number_input(
            "حداکثر P/E پیش‌فرض",
            min_value=0,
            max_value=50,
            value=int(st.session_state.config["DEFAULT_FILTERS"]["max_pe"]),
            step=1,
            key="default_max_pe"
        )
    with col3:
        default_min_volume = st.number_input(
            "حداقل حجم پیش‌فرض",
            min_value=0,
            max_value=50000000,
            value=int(st.session_state.config["DEFAULT_FILTERS"]["min_volume"]),
            step=100000,
            key="default_min_volume"
        )
    
    st.session_state.config["DEFAULT_FILTERS"]["min_quality"] = int(default_min_quality)
    st.session_state.config["DEFAULT_FILTERS"]["max_pe"] = int(default_max_pe)
    st.session_state.config["DEFAULT_FILTERS"]["min_volume"] = int(default_min_volume)
    
    st.success("✅ تنظیمات فیلترهای پیش‌فرض ذخیره شد!")
    
    st.markdown("---")
    st.caption("💡 تمام تغییرات به صورت خودکار ذخیره می‌شوند و روی تحلیل‌ها اعمال می‌گردند.")

# =====================================================
# =============== آمار کلی ============================
# =====================================================

st.subheader("📊 آمار کلی بازار")

col1, col2, col3, col4, col5 = st.columns(5)

with col1:
    st.metric("📊 تعداد کل نمادها", f"{len(filtered_df):,}")
with col2:
    total_mv = filtered_df['mv'].sum() / 10_000_000_000_000
    st.metric("💰 ارزش کل بازار", f"{total_mv:.1f} تریلیون تومان")
with col3:
    pe_avg = filtered_df['pe'].mean()
    st.metric("📈 میانگین P/E", f"{pe_avg:.2f}")
with col4:
    high_quality = len(filtered_df[filtered_df['quality_score'] >= 70])
    st.metric("⭐ سهام با کیفیت بالا", high_quality)
with col5:
    money_inflow = len(filtered_df[filtered_df['money_flow_status'].str.contains('ورود پول', na=False)])
    st.metric("🟢 نمادهای با ورود پول", money_inflow)

st.markdown("---")

# =====================================================
# =============== تب 1: جدول تحلیل ====================
# =====================================================

with tab1:
    st.subheader("📋 جدول تحلیل بنیادی")
    
    default_filters = config["DEFAULT_FILTERS"]
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        min_quality = st.slider(
            "حداقل کیفیت",
            0, 100,
            default_filters["min_quality"],
            key="filter_min_quality"
        )
    with col2:
        max_pe = st.slider(
            "حداکثر P/E",
            0, 50,
            default_filters["max_pe"],
            key="filter_max_pe"
        )
    with col3:
        min_volume = st.number_input(
            "حداقل حجم",
            min_value=0,
            value=default_filters["min_volume"],
            step=100000,
            key="filter_min_volume"
        )
    with col4:
        money_filter = st.selectbox(
            "جریان پول",
            ['همه', '🟢 ورود پول قوی', '🟡 ورود پول', '🔵 متوسط', '⚪ خروج پول', '⛔ خروج پول قوی'],
            key="filter_money_flow"
        )
    
    # اعمال فیلترها روی filtered_df
    result_df = filtered_df[
        (filtered_df['quality_score'] >= min_quality) &
        (filtered_df['pe'] <= max_pe) &
        (filtered_df['tvol'] >= min_volume)
    ]
    
    if money_filter != 'همه':
        result_df = result_df[result_df['money_flow_status'] == money_filter]
    
    result_df = result_df.sort_values('quality_score', ascending=False)
    
    # ساخت جدول HTML
    html = """
    <table class="custom-table">
    <thead>
    <tr>
        <th>#</th>
        <th>نماد</th>
        <th>نام شرکت</th>
        <th>صنعت</th>
        <th>ارزش بازار</th>
        <th>قیمت</th>
        <th>تغییر %</th>
        <th>P/E</th>
        <th>EPS</th>
        <th>حجم</th>
        <th>کیفیت</th>
        <th>وضعیت</th>
    </tr>
    </thead>
    <tbody>
    """
    
    for i, (_, row) in enumerate(result_df.head(100).iterrows(), 1):
        change = row['plp']
        if change is not None and change > 0:
            change_display = f'<span style="color:#16a34a; font-weight:bold;">▲ +{change:.2f}%</span>'
        elif change is not None and change < 0:
            change_display = f'<span style="color:#dc2626; font-weight:bold;">▼ {change:.2f}%</span>'
        else:
            change_display = f"{change:.2f}%"
        
        pe_val = row['pe']
        if pe_val is not None:
            color = pe_color(pe_val, config)
            pe_display = f'<span style="color:{color}; font-weight:bold;">{pe_label(pe_val)}</span>'
        else:
            pe_display = "—"
        
        quality = row['quality_score']
        if quality >= 70:
            quality_color = "#16a34a"
            quality_star = "⭐"
        elif quality >= 50:
            quality_color = "#ca8a04"
            quality_star = "🌟"
        elif quality >= 30:
            quality_color = "#f97316"
            quality_star = "👍"
        else:
            quality_color = "#dc2626"
            quality_star = "⚠️"
        quality_display = f'<span style="color:{quality_color}; font-weight:bold;">{quality_star} {quality:.0f}</span>'
        
        # صنعت
        industry = row.get('cs', '—') if 'cs' in row else '—'
        
        html += f"""
    <tr>
        <td style="font-weight:bold; color:#0f172a;">{to_persian_number(i)}</td>
        <td style="font-weight:bold; color:#0f172a;">{row['l18']}</td>
        <td style="text-align:right; padding-right:14px;">{row['l30']}</td>
        <td>{industry}</td>
        <td style="font-weight:bold; color:#0f172a;">{fmt_market_value_persian(row['mv'])}</td>
        <td>{fmt(row['pc'])}</td>
        <td>{change_display}</td>
        <td>{pe_display}</td>
        <td>{fmt(row['eps'])}</td>
        <td>{fmt(row['tvol'])}</td>
        <td>{quality_display}</td>
        <td>{row['money_flow_status']}</td>
    </tr>
    """
    
    html += """
    </tbody>
    </table>
    """
    
    st.markdown(html, unsafe_allow_html=True)
    st.caption(f"📌 نمایش {min(100, len(result_df))} از {len(result_df)} نماد | 🟢 مناسب  🟡 متوسط  🔴 گران")

# =====================================================
# =============== تب 2: بهترین‌ها ====================
# =====================================================

with tab2:
    st.subheader("⭐ ۱۰ سهم برتر از نظر کیفیت")
    
    top_quality = filtered_df.nlargest(10, 'quality_score')
    
    fig = px.bar(
        top_quality,
        x='l18',
        y='quality_score',
        color='quality_score',
        color_continuous_scale='Viridis',
        title='۱۰ سهم با بالاترین امتیاز کیفیت',
        labels={'l18': 'نماد', 'quality_score': 'امتیاز کیفیت'}
    )
    fig.update_layout(
        height=400, 
        showlegend=False,
        font=dict(family="Vazirmatn, Tahoma, sans-serif", size=12),
        title_font=dict(family="Vazirmatn, Tahoma, sans-serif", size=14)
    )
    fig.update_traces(texttemplate='%{y:.0f}', textposition='outside')
    st.plotly_chart(fig, use_container_width=True)
    
    st.markdown("### 📋 جزئیات سهام برتر")
    
    for _, row in top_quality.iterrows():
        with st.expander(f"⭐ {row['l18']} - {row['l30']}"):
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("🏅 امتیاز کیفیت", f"{row['quality_score']:.0f}/100")
                st.metric("📊 P/E", f"{row['pe']:.2f}")
            with col2:
                st.metric("💰 EPS", f"{row['eps']:,.0f}")
                st.metric("🏢 ارزش بازار", f"{row['mv_billion']:.1f} میلیارد")
            with col3:
                st.metric("💵 جریان پول", row['money_flow_status'])
                st.metric("📈 تغییر قیمت", f"{row['plp']:.2f}%")
            with col4:
                st.metric("📊 حجم معاملات", f"{row['tvol']:,.0f}")
                st.metric("🎯 رتبه کلی", f"#{int(row['overall_rank'])}")

# =====================================================
# =============== تب 3: جریان پول ====================
# =====================================================

with tab3:
    st.subheader("💰 تحلیل جریان پول")
    
    col1, col2 = st.columns(2)
    
    with col1:
        money_dist = filtered_df['money_flow_status'].value_counts()
        fig = px.pie(
            values=money_dist.values,
            names=money_dist.index,
            title='توزیع وضعیت جریان پول',
            color_discrete_sequence=px.colors.qualitative.Set3
        )
        fig.update_traces(textposition='inside', textinfo='percent+label')
        fig.update_layout(
            font=dict(family="Vazirmatn, Tahoma, sans-serif", size=12),
            title_font=dict(family="Vazirmatn, Tahoma, sans-serif", size=14)
        )
        st.plotly_chart(fig, use_container_width=True)
    
    with col2:
        top_money = filtered_df.nlargest(20, 'money_flow')
        fig = px.bar(
            top_money,
            x='l18',
            y='money_flow',
            color='money_flow_status',
            title='۲۰ نماد با بیشترین جریان پول',
            labels={'l18': 'نماد', 'money_flow': 'جریان پول (میلیارد ریال)'}
        )
        fig.update_layout(
            height=400,
            font=dict(family="Vazirmatn, Tahoma, sans-serif", size=12),
            title_font=dict(family="Vazirmatn, Tahoma, sans-serif", size=14)
        )
        st.plotly_chart(fig, use_container_width=True)

# =====================================================
# =============== تب 4: تحلیل حجم ====================
# =====================================================

with tab4:
    st.subheader("📈 تحلیل حجم معاملات")
    
    col1, col2 = st.columns(2)
    with col1:
        top_n = st.slider("تعداد نمادهای برتر", 5, 50, 20)
    with col2:
        sort_by = st.selectbox("مرتب‌سازی بر اساس", ['حجم', 'ارزش بازار', 'تغییر قیمت'])
    
    sorted_df = filtered_df.nlargest(top_n, 'tvol')
    
    fig = go.Figure()
    fig.add_trace(go.Bar(
        x=sorted_df['l18'],
        y=sorted_df['tvol'],
        name='حجم معاملات',
        marker_color='lightblue',
        text=sorted_df['tvol'].apply(lambda x: f'{x/1_000_000:.1f}M'),
        textposition='outside'
    ))
    fig.add_trace(go.Scatter(
        x=sorted_df['l18'],
        y=sorted_df['pc'],
        name='قیمت',
        yaxis='y2',
        line=dict(color='red', width=2),
        mode='lines+markers'
    ))
    fig.update_layout(
        title=f'برترین نمادها از نظر حجم معاملات',
        yaxis=dict(title='حجم معاملات'),
        yaxis2=dict(title='قیمت', overlaying='y', side='right'),
        height=450,
        hovermode='x unified',
        font=dict(family="Vazirmatn, Tahoma, sans-serif", size=12),
        title_font=dict(family="Vazirmatn, Tahoma, sans-serif", size=14)
    )
    st.plotly_chart(fig, use_container_width=True)
    
    st.subheader("📊 آمار حجم معاملات")
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("میانگین حجم", f"{filtered_df['tvol'].mean():,.0f}")
    with col2:
        st.metric("میانگین حجم (میلیون)", f"{filtered_df['tvol'].mean()/1_000_000:.1f}M")
    with col3:
        st.metric("بیشترین حجم", f"{filtered_df['tvol'].max():,.0f}")
    with col4:
        st.metric("کمترین حجم", f"{filtered_df['tvol'].min():,.0f}")

# =====================================================
# =============== تب 5: خروجی اکسل ====================
# =====================================================

with tab5:
    st.subheader("📥 خروجی اکسل")
    st.markdown("داده‌های تحلیل شده را به صورت فایل اکسل (Excel) فرمت‌بندی‌شده دانلود کنید.")
    
    def build_excel(df_export):
        wb = Workbook()
        ws = wb.active
        ws.title = "تحلیل بازار"
        ws.sheet_view.rightToLeft = True
        
        headers = [
            "نماد", "نام شرکت", "صنعت", "ارزش بازار (میلیارد)", "قیمت", "تغییر %",
            "P/E", "EPS", "حجم معاملات", "جریان پول", "امتیاز کیفیت",
            "برچسب کیفیت", "وضعیت جریان پول", "رتبه کلی"
        ]
        
        header_font = Font(bold=True, color="FFFFFF", size=11, name="Vazirmatn")
        header_fill = PatternFill("solid", fgColor="0F172A")
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_align = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )
        green_fill = PatternFill("solid", fgColor="DCFCE7")
        yellow_fill = PatternFill("solid", fgColor="FEF9C3")
        red_fill = PatternFill("solid", fgColor="FEE2E2")
        alt_fill = PatternFill("solid", fgColor="F8FAFC")
        
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin
        
        ws.row_dimensions[1].height = 28
        
        for r_idx, (_, row) in enumerate(df_export.iterrows(), 2):
            values = [
                row['l18'],
                row['l30'],
                row.get('cs', '—'),
                row['mv_billion'],
                row['pc'],
                row['plp'],
                row['pe'],
                row['eps'],
                row['tvol'],
                row['money_flow'],
                row['quality_score'],
                row['quality_label'],
                row['money_flow_status'],
                row['overall_rank']
            ]
            
            for c_idx, val in enumerate(values, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.alignment = cell_align
                cell.border = thin
                if r_idx % 2 == 0:
                    cell.fill = alt_fill
                
                if c_idx in (1, 2, 3, 12, 13):
                    cell.value = val if val is not None else "—"
                    if c_idx == 13:
                        if "ورود پول قوی" in str(val):
                            cell.font = Font(color="16A34A", bold=True)
                        elif "ورود پول" in str(val):
                            cell.font = Font(color="CA8A04", bold=True)
                        elif "خروج پول" in str(val):
                            cell.font = Font(color="DC2626", bold=True)
                    continue
                
                if val is None or pd.isna(val):
                    cell.value = "—"
                else:
                    cell.value = float(val)
                    
                    if c_idx == 4:
                        cell.number_format = '#,##0.0'
                    elif c_idx in (5, 8, 9, 10):
                        cell.number_format = '#,##0'
                    elif c_idx == 6:
                        cell.number_format = '0.00"%"'
                        if val > 0:
                            cell.font = Font(color="16A34A")
                        elif val < 0:
                            cell.font = Font(color="DC2626")
                    elif c_idx == 7:
                        cell.number_format = '0.00'
                        if val < config["PE_COLORS"]["excellent"]:
                            cell.fill = green_fill
                        elif val <= config["PE_COLORS"]["good"]:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    elif c_idx == 11:
                        cell.number_format = '0'
                        if val >= 70:
                            cell.fill = green_fill
                        elif val >= 50:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill
                    elif c_idx == 14:
                        cell.number_format = '0'
        
        widths = [12, 22, 18, 18, 14, 12, 12, 14, 16, 18, 14, 16, 20, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        
        ws.freeze_panes = "A2"
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
    
    try:
        excel_buffer = build_excel(filtered_df)
        st.download_button(
            label="📥 دانلود فایل اکسل فرمت‌بندی‌شده",
            data=excel_buffer,
            file_name=f"تحلیل_بازار_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.success("✅ فایل اکسل با موفقیت ایجاد شد!")
    except Exception as e:
        st.error(f"❌ خطا در ایجاد فایل اکسل: {e}")
        st.info("💡 در حال استفاده از روش جایگزین (CSV)...")
        
        csv = filtered_df.to_csv(index=False, encoding='utf-8-sig')
        st.download_button(
            label="📥 دانلود CSV (جایگزین)",
            data=csv.encode('utf-8-sig'),
            file_name=f"تحلیل_بازار_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
            mime="text/csv"
        )
    
    with st.expander("📋 مشاهده پیش‌نمایش داده‌ها"):
        st.dataframe(
            filtered_df[['l18', 'l30', 'quality_score', 'quality_label', 'money_flow_status', 
               'pe', 'eps', 'tvol', 'plp']].head(10).rename(columns={
                   'l18': 'نماد',
                   'l30': 'نام شرکت',
                   'quality_score': 'امتیاز کیفیت',
                   'quality_label': 'برچسب کیفیت',
                   'money_flow_status': 'وضعیت جریان پول',
                   'pe': 'P/E',
                   'eps': 'EPS',
                   'tvol': 'حجم معاملات',
                   'plp': 'تغییر %'
               }),
            height=300
        )

# =====================================================
# =============== فوتر ================================
# =====================================================

st.markdown("---")

st.markdown("""
<div class="footer-text">
    <span class="highlight">📊 تحلیل پیشرفته بازار</span>
    <span class="divider">|</span>
    <span>تحلیلی از <strong style="color:#0f172a;">داود شورگشتی</strong></span>
    <span class="divider">|</span>
    <span style="color:#94a3b8;">📅 {}</span>
</div>
""".format(datetime.now().strftime('%Y/%m/%d - %H:%M')), unsafe_allow_html=True)