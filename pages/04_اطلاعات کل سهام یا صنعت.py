import streamlit as st
import sys
import os
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
import jdatetime
import math

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from utils.database import get_connection, init_db

st.set_page_config(page_title="مقایسه شرکت‌ها", layout="wide")
from utils.styles import apply_styles
apply_styles()


st.markdown("""
<style>
    .main .block-container { direction: rtl; padding-top: 1.2rem; }
    .stTable, table { direction: rtl; text-align: center; }
    h1 { font-size: 1.6rem !important; }
    h2, h3 { font-size: 1.2rem !important; }
    .progress-box {
        display: flex;
        gap: 4px;
        justify-content: center;
        align-items: center;
    }
    .progress-square {
        width: 18px;
        height: 18px;
        border-radius: 3px;
        border: 1px solid #cbd5e1;
        display: inline-block;
    }
    .progress-square.filled {
        background: #2563eb;
        border-color: #2563eb;
    }
    .progress-square.empty {
        background: #e8ecf1;
    }
    .progress-label {
        font-size: 11px;
        color: #64748b;
        margin-right: 4px;
    }
    .coverage-good { color: #16a34a; font-weight: bold; }
    .coverage-mid { color: #ca8a04; font-weight: bold; }
    .coverage-bad { color: #dc2626; font-weight: bold; }
    .status-ahead { color: #16a34a; font-weight: bold; }
    .status-behind { color: #dc2626; font-weight: bold; }
    .status-on-track { color: #ca8a04; font-weight: bold; }
    .export-yes { color: #16a34a; font-weight: bold; font-size: 18px; }
    .export-no { color: #94a3b8; font-size: 18px; }
</style>
""", unsafe_allow_html=True)

st.title("مقایسه شرکت‌ها")

init_db()

COLORS = ["#2563eb", "#dc2626", "#16a34a", "#ea580c", "#7c3aed", "#0891b2", "#db2777", "#0d9488"]


# ============================================================
# سال مالی
# ============================================================
def fiscal_period_end(fiscal_year, period_type, fiscal_end_month, fiscal_end_day=29):
    months_back = {12: 0, 9: 3, 6: 6, 3: 9}
    back = months_back.get(int(period_type), 0)
    end_month = int(fiscal_end_month) - back
    cal_year = int(fiscal_year)
    if end_month <= 0:
        end_month += 12
        cal_year -= 1
    return cal_year, end_month, int(fiscal_end_day or 29)


def resolve_fiscal_year(row, fiscal_end_month, fiscal_end_day=29):
    y = int(row["year_solar"])
    pt = int(row["period_type"])
    F = int(fiscal_end_month)
    day = int(fiscal_end_day or 29)

    end_m = None
    try:
        if row["end_month"] is not None:
            end_m = int(row["end_month"])
    except Exception:
        end_m = None

    candidates = [y, y + 1, y - 1, y + 2, y - 2]

    if end_m:
        for cand in candidates:
            cy, em, _ = fiscal_period_end(cand, pt, F, day)
            if em == end_m and cy == y:
                return cand
        for cand in candidates:
            cy, em, _ = fiscal_period_end(cand, pt, F, day)
            if em == end_m and cand == y:
                return cand
        for cand in candidates:
            cy, em, _ = fiscal_period_end(cand, pt, F, day)
            if em == end_m:
                return cand
    return y


def get_first_quarter_months(fiscal_end_month):
    start = (int(fiscal_end_month) % 12) + 1
    return [((start - 1 + i) % 12) + 1 for i in range(3)]


def calendar_year_for_fiscal_month(fiscal_year, month, fiscal_end_month):
    if int(month) > int(fiscal_end_month):
        return int(fiscal_year) - 1
    return int(fiscal_year)


def get_q1_month_year_pairs(fiscal_year, fiscal_end_month):
    pairs = []
    for m in get_first_quarter_months(fiscal_end_month):
        cy = calendar_year_for_fiscal_month(fiscal_year, m, fiscal_end_month)
        pairs.append((cy, m))
    return pairs


def get_progress_based_on_reports(monthly_rows, target_year, fiscal_end_month):
    """
    محاسبه پیشرفت بر اساس تعداد ماه‌هایی که گزارش فروش برای آنها ثبت شده
    """
    try:
        start_month = (fiscal_end_month % 12) + 1
        reported_months = 0
        total_sales = 0
        has_export = False
        
        for i in range(12):
            m = ((start_month - 1 + i) % 12) + 1
            cy = calendar_year_for_fiscal_month(target_year, m, fiscal_end_month)
            
            found = False
            for r in monthly_rows:
                if int(r["year_solar"]) == cy and int(r["month"]) == m:
                    tot = r["total_sales"]
                    if tot is None:
                        tot = (r["domestic_sales"] or 0) + (r["export_sales"] or 0)
                    if tot and float(tot) > 0:
                        found = True
                        total_sales += float(tot)
                        # بررسی فروش صادراتی
                        if r["export_sales"] and float(r["export_sales"]) > 0:
                            has_export = True
                    break
            
            if found:
                reported_months += 1
            else:
                break
        
        progress = (reported_months / 12) * 100
        return min(100, progress), reported_months, total_sales, has_export
    except Exception:
        return None, 0, 0, False


# ============================================================
# دیتابیس
# ============================================================
def get_companies():
    conn = get_connection()
    rows = conn.execute("""
        SELECT id, symbol, name_fa, industry, market_value, rank_in_industry,
               fiscal_end_month, fiscal_end_day
        FROM companies ORDER BY symbol
    """).fetchall()
    conn.close()
    return rows


def get_periods(company_id):
    conn = get_connection()
    rows = conn.execute("""
        SELECT p.id, p.year_solar, p.period_type, p.end_month, p.end_day,
               f.operating_revenue, f.cogs, f.other_income, f.non_operating_income,
               f.net_profit, f.comprehensive_income, f.inventory,
               f.trade_receivables, f.equity, f.current_assets,
               f.total_assets, f.approved_dividend
        FROM periods p
        LEFT JOIN financials f ON f.period_id = p.id
        WHERE p.company_id = ?
        ORDER BY p.year_solar DESC, p.period_type DESC
    """, (company_id,)).fetchall()
    conn.close()
    return rows


def get_all_monthly_sales(company_id):
    conn = get_connection()
    rows = conn.execute("""
        SELECT year_solar, month, domestic_sales, export_sales, total_sales
        FROM monthly_sales WHERE company_id = ?
        ORDER BY year_solar, month
    """, (company_id,)).fetchall()
    conn.close()
    return rows


def get_recurring_by_year(company_id):
    conn = get_connection()
    rows = conn.execute("""
        SELECT year_solar, SUM(amount) as total
        FROM non_operating_items
        WHERE company_id = ? AND is_recurring = 1
        GROUP BY year_solar ORDER BY year_solar
    """, (company_id,)).fetchall()
    conn.close()
    return {r["year_solar"]: float(r["total"] or 0) for r in rows}


# ============================================================
# محاسبات
# ============================================================
def safe_div(a, b):
    try:
        if a is None or b is None or float(b) == 0:
            return None
        return float(a) / float(b)
    except Exception:
        return None


def positive_only(v):
    try:
        if v is None:
            return 0.0
        x = float(v)
        return x if x > 0 else 0.0
    except Exception:
        return 0.0


def calc_metrics(row):
    revenue = row["operating_revenue"]
    other_inc = positive_only(row["other_income"])
    non_op = positive_only(row["non_operating_income"])
    net_profit = row["net_profit"]

    op_profit = None
    if net_profit is not None:
        op_profit = float(net_profit) - other_inc - non_op

    net_margin = safe_div(op_profit, revenue)
    recv_ratio = safe_div(row["trade_receivables"], row["current_assets"])
    div_ratio = safe_div(row["approved_dividend"], row["comprehensive_income"])
    return op_profit, net_margin, recv_ratio, div_ratio


def calc_forward_estimates(company, periods, monthly_rows):
    result = {
        "has_data": False,
        "sales_final": None, "margin_avg": None,
        "est_profit": None, "est_recurring_non_op": None,
        "est_dividend": None, "payout_avg": None,
        "pe_forward": None, "pd_forward": None, "ps_forward": None,
        "margin_annual": None, "margin_last": None,
        "latest_revenue": None, "latest_period_type": None,
        "progress": None,
        "reported_months": 0,
        "actual_sales_current_year": None,
        "has_export": False,
    }

    if not periods and not monthly_rows:
        return result

    fiscal_end_month = int(company["fiscal_end_month"] or 12)
    fiscal_end_day = int(company["fiscal_end_day"] or 29)
    market_value = company["market_value"]

    fiscal_years = set()
    for r in periods:
        fiscal_years.add(resolve_fiscal_year(r, fiscal_end_month, fiscal_end_day))
    for r in monthly_rows:
        fiscal_years.add(int(r["year_solar"]))

    all_years = sorted(y for y in fiscal_years if 1390 <= y <= 1410)
    if not all_years:
        return result

    last_year = max(all_years)

    def has_annual_for(fy):
        for r in periods:
            if resolve_fiscal_year(r, fiscal_end_month, fiscal_end_day) == fy and int(r["period_type"]) == 12:
                return True
        return False

    has_annual_last = has_annual_for(last_year)
    target_year = last_year + 1 if has_annual_last else last_year

    # ============================================================
    # پیشرفت بر اساس گزارش‌های موجود
    # ============================================================
    progress, reported_months, actual_sales, has_export = get_progress_based_on_reports(
        monthly_rows, target_year, fiscal_end_month
    )
    result["progress"] = progress
    result["reported_months"] = reported_months
    result["actual_sales_current_year"] = actual_sales if actual_sales > 0 else None
    result["has_export"] = has_export

    # ============================================================
    # دریافت آخرین دوره مالی
    # ============================================================
    latest_revenue = None
    latest_period_type = None
    if periods:
        enriched = []
        for r in periods:
            fy = resolve_fiscal_year(r, fiscal_end_month, fiscal_end_day)
            enriched.append((fy, int(r["period_type"]), r))
        enriched.sort(key=lambda x: (x[0], x[1]), reverse=True)
        if enriched:
            latest_revenue = enriched[0][2]["operating_revenue"]
            latest_period_type = enriched[0][1]
    result["latest_revenue"] = latest_revenue
    result["latest_period_type"] = latest_period_type

    # ============================================================
    # برآورد فروش
    # ============================================================
    q1 = None
    for r in periods:
        if resolve_fiscal_year(r, fiscal_end_month, fiscal_end_day) == target_year and int(r["period_type"]) == 3:
            q1 = r
            break

    monthly_index = {}
    for r in monthly_rows:
        tot = r["total_sales"]
        if tot is None:
            tot = (r["domestic_sales"] or 0) + (r["export_sales"] or 0)
        if tot and float(tot) > 0:
            monthly_index[(int(r["year_solar"]), int(r["month"]))] = float(tot)

    q1_pairs = get_q1_month_year_pairs(target_year, fiscal_end_month)

    def collect_q1_sales():
        month_sales = {}
        for cy, m in q1_pairs:
            val = monthly_index.get((cy, m))
            if val is not None and val > 0:
                month_sales[m] = val
        return month_sales

    sales_m1 = sales_m2 = None

    if q1 and q1["operating_revenue"] and float(q1["operating_revenue"]) > 0:
        rev = float(q1["operating_revenue"])
        sales_m1 = (rev / 3.0) * 12.0
        month_sales = collect_q1_sales()
        if month_sales:
            ordered = [m for m in get_first_quarter_months(fiscal_end_month) if m in month_sales]
            last_val = month_sales[ordered[-1]]
            sum_existing = sum(month_sales.values())
            sales_m2 = last_val * (12 - len(month_sales)) + sum_existing
        else:
            sales_m2 = sales_m1
    else:
        month_sales = collect_q1_sales()
        if not month_sales:
            return result
        n = len(month_sales)
        avg_month = sum(month_sales.values()) / float(n)
        sales_m1 = avg_month * 12.0
        ordered = [m for m in get_first_quarter_months(fiscal_end_month) if m in month_sales]
        last_val = month_sales[ordered[-1]]
        sum_existing = sum(month_sales.values())
        sales_m2 = last_val * (12 - n) + sum_existing

    sales_final = min(sales_m1, sales_m2) if sales_m1 and sales_m2 else (sales_m1 or sales_m2)
    result["sales_final"] = sales_final

    # ============================================================
    # حاشیه سود
    # ============================================================
    margin_annual = None
    for r in periods:
        if resolve_fiscal_year(r, fiscal_end_month, fiscal_end_day) == target_year - 1 and int(r["period_type"]) == 12:
            _, margin, _, _ = calc_metrics(r)
            if margin is not None:
                margin_annual = margin
            break

    margin_last = None
    candidates = []
    for r in periods:
        fy = resolve_fiscal_year(r, fiscal_end_month, fiscal_end_day)
        if fy <= target_year:
            candidates.append((fy, int(r["period_type"]), r))
    if candidates:
        non_annual = [c for c in candidates if c[1] != 12]
        pool = non_annual if non_annual else candidates
        last = max(pool, key=lambda x: (x[0], x[1]))
        _, margin, _, _ = calc_metrics(last[2])
        if margin is not None:
            margin_last = margin

    result["margin_annual"] = margin_annual
    result["margin_last"] = margin_last

    if margin_annual is not None and margin_last is not None:
        margin_avg = (margin_annual + margin_last) / 2.0
    elif margin_annual is not None:
        margin_avg = margin_annual
    elif margin_last is not None:
        margin_avg = margin_last
    else:
        return result

    result["margin_avg"] = margin_avg
    est_profit = sales_final * margin_avg

    # ============================================================
    # اقلام غیرعملیاتی
    # ============================================================
    recurring_map = get_recurring_by_year(company["id"])
    relevant = [y for y in sorted(recurring_map.keys()) if y < target_year]
    est_recurring = None
    if len(relevant) >= 2:
        v1, v2 = recurring_map[relevant[-2]], recurring_map[relevant[-1]]
        g = (v2 - v1) / v1 if v1 != 0 else 0
        est_recurring = v2 * (1 + g)
    elif len(relevant) == 1:
        est_recurring = recurring_map[relevant[-1]]

    result["est_recurring_non_op"] = est_recurring
    if est_recurring is not None:
        est_profit = est_profit + est_recurring
    result["est_profit"] = est_profit

    # ============================================================
    # نسبت سود تقسیمی
    # ============================================================
    payouts = []
    for y in [target_year - 1, target_year - 2]:
        for r in periods:
            if resolve_fiscal_year(r, fiscal_end_month, fiscal_end_day) == y and int(r["period_type"]) == 12:
                div, comp = r["approved_dividend"], r["comprehensive_income"]
                if div is not None and comp is not None and float(comp) != 0:
                    payouts.append(float(div) / float(comp))
                break
    payout_avg = sum(payouts) / len(payouts) if payouts else None
    result["payout_avg"] = payout_avg
    if payout_avg is not None and est_profit is not None:
        result["est_dividend"] = est_profit * payout_avg

    # ============================================================
    # نسبت‌های Forward
    # ============================================================
    if market_value and est_profit and est_profit > 0:
        result["pe_forward"] = float(market_value) / est_profit
    if market_value and result.get("est_dividend") and result["est_dividend"] > 0:
        result["pd_forward"] = float(market_value) / result["est_dividend"]
    if market_value and sales_final and sales_final > 0:
        result["ps_forward"] = float(market_value) / sales_final

    result["has_data"] = True
    return result


def fmt(v):
    if v is None:
        return "—"
    try:
        return f"{float(v):,.0f}"
    except Exception:
        return "—"


def fmt_pct(v):
    if v is None:
        return "—"
    try:
        return f"{float(v)*100:.1f}%"
    except Exception:
        return "—"


def fmt_ratio(v):
    if v is None:
        return "—"
    try:
        return f"{float(v):.2f}"
    except Exception:
        return "—"


def pe_color(pe):
    if pe is None:
        return "#6b7280"
    if pe < 5:
        return "#16a34a"
    if pe <= 7:
        return "#ca8a04"
    return "#dc2626"


def pe_label(pe):
    if pe is None:
        return "—"
    if pe < 5:
        return f"{pe:.2f} (مناسب)"
    if pe <= 7:
        return f"{pe:.2f} (متوسط)"
    return f"{pe:.2f} (بالا)"


def pe_html(pe):
    if pe is None:
        return "—"
    color = pe_color(pe)
    label = pe_label(pe)
    return f'<span style="color:{color}; font-weight:bold;">{label}</span>'


def progress_html(progress, reported_months=None):
    """نمایش پیشرفت با ۴ مربع و تعداد ماه‌های گزارش‌شده"""
    if progress is None:
        return "—"
    
    try:
        progress_val = float(progress)
    except (TypeError, ValueError):
        return "—"
    
    if math.isnan(progress_val):
        return "—"
    
    if progress_val < 0 or progress_val > 100:
        return "—"
    
    filled = int(round(progress_val / 25))
    filled = min(4, max(0, filled))
    
    squares = []
    for i in range(4):
        if i < filled:
            squares.append('<span class="progress-square filled"></span>')
        else:
            squares.append('<span class="progress-square empty"></span>')
    
    month_text = ""
    if reported_months is not None and reported_months > 0:
        month_text = f'<span style="font-size:10px; color:#64748b; margin-right:4px;">({reported_months} ماه)</span>'
    
    return f'<div class="progress-box"><span class="progress-label">{int(progress_val)}%</span>{"".join(squares)}{month_text}</div>'


# ============================================================
# ✅ تابع اصلاح‌شده نهایی get_status_html
# ============================================================
def get_status_html(coverage, progress):
    """بررسی وضعیت نسبت به برنامه - اصلاح‌شده نهایی"""
    if coverage is None or progress is None:
        return '<span style="color:#6b7280;">—</span>'
    
    try:
        coverage_val = float(coverage)
        progress_val = float(progress)
    except:
        return '<span style="color:#6b7280;">—</span>'
    
    # coverage به صورت اعشار هست (مثلاً 1.241)
    # progress به صورت درصد هست (مثلاً 58)
    progress_val = progress_val / 100
    
    # اگر coverage بیشتر از 10 باشه، یعنی به صورت درصد ذخیره شده (مثلاً 124)
    if coverage_val > 10:
        coverage_val = coverage_val / 100
    
    diff = coverage_val - progress_val
    diff_percent = diff * 100
    
    if diff > 0.05:
        return f'<span class="status-ahead">✅ جلوتر ({diff_percent:.1f}%)</span>'
    elif diff < -0.05:
        return f'<span class="status-behind">⚠️ عقب‌تر ({abs(diff_percent):.1f}%)</span>'
    else:
        return f'<span class="status-on-track">📊 طبق برنامه</span>'


def get_export_html(has_export):
    """نمایش وضعیت صادراتی"""
    if has_export:
        return '<span class="export-yes">✅</span>'
    else:
        return '<span class="export-no">—</span>'


# ============================================================
# انتخاب شرکت‌ها
# ============================================================
companies = get_companies()
if not companies:
    st.warning("هنوز شرکتی ثبت نشده است.")
    st.stop()

industries = sorted(set(c["industry"] for c in companies if c["industry"]))

col_f1, col_f2 = st.columns(2)
with col_f1:
    mode = st.radio(
        "نحوه انتخاب:",
        options=["انتخاب دستی", "همه شرکت‌ها", "بر اساس صنعت"],
        horizontal=True
    )
with col_f2:
    selected_industry = None
    if mode == "بر اساس صنعت":
        if industries:
            selected_industry = st.selectbox("صنعت:", options=industries)
        else:
            st.warning("صنعتی ثبت نشده است.")
            st.stop()

options = {f"{c['symbol']} — {c['name_fa'] or ''}": c for c in companies}

if mode == "همه شرکت‌ها":
    selected_companies = list(companies)
elif mode == "بر اساس صنعت":
    selected_companies = [c for c in companies if c["industry"] == selected_industry]
    if not selected_companies:
        st.warning("شرکتی در این صنعت نیست.")
        st.stop()
else:
    selected_labels = st.multiselect(
        "شرکت‌های مورد نظر (حداقل ۲):",
        options=list(options.keys()),
        default=list(options.keys())[: min(3, len(options))]
    )
    if len(selected_labels) < 1:
        st.info("حداقل یک شرکت انتخاب کنید.")
        st.stop()
    selected_companies = [options[lb] for lb in selected_labels]

st.caption(f"تعداد شرکت‌های انتخاب‌شده: **{len(selected_companies)}**")

# ============================================================
# محاسبه داده
# ============================================================
rows_data = []
for comp in selected_companies:
    periods = get_periods(comp["id"])
    monthly = get_all_monthly_sales(comp["id"])
    fwd = calc_forward_estimates(comp, periods, monthly)
    latest_margin = latest_recv = None
    if periods:
        _, latest_margin, latest_recv, _ = calc_metrics(periods[0])
    
    # محاسبه نسبت پوشش بر اساس فروش محقق‌شده سال جاری
    coverage_ratio = None
    actual_sales = fwd.get("actual_sales_current_year")
    estimated_sales = fwd.get("sales_final")
    
    if actual_sales is not None and estimated_sales and estimated_sales > 0:
        try:
            coverage_ratio = float(actual_sales) / float(estimated_sales)
        except (TypeError, ValueError):
            coverage_ratio = None
    
    rows_data.append({
        "نماد": comp["symbol"],
        "نام": comp["name_fa"] or "—",
        "صنعت": comp["industry"] or "—",
        "ارزش بازار": comp["market_value"],
        "برآورد فروش": fwd.get("sales_final"),
        "فروش محقق‌شده": fwd.get("actual_sales_current_year"),
        "نسبت پوشش برآورد": coverage_ratio,
        "میانگین حاشیه": fwd.get("margin_avg"),
        "برآورد سود": fwd.get("est_profit"),
        "برآورد سود تقسیمی": fwd.get("est_dividend"),
        "P/E Forward": fwd.get("pe_forward"),
        "P/D Forward": fwd.get("pd_forward"),
        "P/S Forward": fwd.get("ps_forward"),
        "حاشیه آخرین دوره": latest_margin,
        "نسبت مطالبات": latest_recv,
        "پیشرفت سال مالی": fwd.get("progress"),
        "تعداد ماه‌های گزارش": fwd.get("reported_months"),
        "صادراتی": fwd.get("has_export"),
    })

df = pd.DataFrame(rows_data)

# ============================================================
# جدول مقایسه + مرتب‌سازی
# ============================================================
st.subheader("جدول مقایسه")

if "sort_by" not in st.session_state:
    st.session_state.sort_by = "P/E Forward"
    st.session_state.sort_asc = True


def safe_sort(dataframe, column, ascending=True):
    temp = dataframe.copy()
    values = []
    for v in temp[column]:
        if v is None or pd.isna(v):
            values.append(1e18 if ascending else -1e18)
        else:
            try:
                values.append(float(v))
            except Exception:
                values.append(1e18 if ascending else -1e18)
    temp["_sort_key"] = values
    temp = temp.sort_values("_sort_key", ascending=ascending).reset_index(drop=True)
    return temp.drop(columns=["_sort_key"])


sort_options = {
    "P/E Forward": "P/E Forward",
    "برآورد سود": "برآورد سود",
    "ارزش بازار": "ارزش بازار",
    "میانگین حاشیه": "میانگین حاشیه",
    "برآورد فروش": "برآورد فروش",
    "P/S Forward": "P/S Forward",
    "P/D Forward": "P/D Forward",
    "نسبت پوشش برآورد": "نسبت پوشش برآورد",
}

s1, s2, s3 = st.columns([3, 2, 5])
with s1:
    chosen = st.selectbox(
        "مرتب‌سازی بر اساس:",
        options=list(sort_options.keys()),
        index=list(sort_options.keys()).index(
            st.session_state.sort_by if st.session_state.sort_by in sort_options else "P/E Forward"
        )
    )
    st.session_state.sort_by = sort_options[chosen]
with s2:
    direction = st.radio(
        "جهت:",
        options=["صعودی ▲", "نزولی ▼"],
        index=0 if st.session_state.sort_asc else 1,
        horizontal=True
    )
    st.session_state.sort_asc = direction.startswith("صعودی")

df_view = safe_sort(df, st.session_state.sort_by, st.session_state.sort_asc)

# ---- جدول HTML ----
html = """
<table style="width:100%; border-collapse:collapse; direction:rtl; text-align:center; font-size:13px;">
<thead>
<tr style="background:#1e3a5f; color:white;">
    <th style="padding:8px; border:1px solid #ccc;">نماد</th>
    <th style="padding:8px; border:1px solid #ccc;">نام</th>
    <th style="padding:8px; border:1px solid #ccc;">صنعت</th>
    <th style="padding:8px; border:1px solid #ccc;" title="شرکت صادراتی دارد؟">صادراتی</th>
    <th style="padding:8px; border:1px solid #ccc;">ارزش بازار</th>
    <th style="padding:8px; border:1px solid #ccc;" title="پیشرفت بر اساس تعداد ماه‌های گزارش‌شده">پیشرفت</th>
    <th style="padding:8px; border:1px solid #ccc;">برآورد فروش</th>
    <th style="padding:8px; border:1px solid #ccc;" title="مجموع فروش ماهانه ثبت‌شده برای سال مالی جاری">فروش محقق‌شده</th>
    <th style="padding:8px; border:1px solid #ccc;" title="نسبت پوشش برآورد">نسبت پوشش</th>
    <th style="padding:8px; border:1px solid #ccc;" title="وضعیت نسبت به برنامه زمانی">وضعیت</th>
    <th style="padding:8px; border:1px solid #ccc;">میانگین حاشیه</th>
    <th style="padding:8px; border:1px solid #ccc;">برآورد سود</th>
    <th style="padding:8px; border:1px solid #ccc;">سود تقسیمی</th>
    <th style="padding:8px; border:1px solid #ccc;">P/E Forward</th>
    <th style="padding:8px; border:1px solid #ccc;">P/D Forward</th>
    <th style="padding:8px; border:1px solid #ccc;">P/S Forward</th>
    <th style="padding:8px; border:1px solid #ccc;">حاشیه آخرین دوره</th>
    <th style="padding:8px; border:1px solid #ccc;">نسبت مطالبات</th>
</tr>
</thead>
<tbody>
"""

for i, (_, row) in enumerate(df_view.iterrows()):
    bg = "#f8fafc" if i % 2 == 0 else "#ffffff"
    
    # P/E
    pe_cell = pe_html(row["P/E Forward"])
    
    # نسبت پوشش
    coverage_val = row["نسبت پوشش برآورد"]
    progress_val = row["پیشرفت سال مالی"]
    reported_months = row["تعداد ماه‌های گزارش"]
    
    if coverage_val is None or pd.isna(coverage_val):
        coverage_display = "—"
        coverage_class = ""
    else:
        coverage_display = f"{coverage_val*100:.1f}%"
        if coverage_val >= 0.9:
            coverage_class = 'coverage-good'
        elif coverage_val >= 0.6:
            coverage_class = 'coverage-mid'
        else:
            coverage_class = 'coverage-bad'
    
    # وضعیت - با تابع اصلاح‌شده نهایی
    status_display = get_status_html(coverage_val, progress_val)
    
    # پیشرفت با تعداد ماه‌ها
    progress_display = progress_html(progress_val, reported_months)
    
    # صادراتی
    export_display = get_export_html(row["صادراتی"])
    
    html += f"""
<tr style="background:{bg};">
    <td style="padding:7px; border:1px solid #ddd; font-weight:bold;">{row['نماد']}</td>
    <td style="padding:7px; border:1px solid #ddd;">{row['نام']}</td>
    <td style="padding:7px; border:1px solid #ddd;">{row['صنعت']}</td>
    <td style="padding:7px; border:1px solid #ddd;">{export_display}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['ارزش بازار'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{progress_display}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['برآورد فروش'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['فروش محقق‌شده'])}</td>
    <td style="padding:7px; border:1px solid #ddd;" class="{coverage_class}">{coverage_display}</td>
    <td style="padding:7px; border:1px solid #ddd;">{status_display}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_pct(row['میانگین حاشیه'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['برآورد سود'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt(row['برآورد سود تقسیمی'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{pe_cell}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_ratio(row['P/D Forward'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_ratio(row['P/S Forward'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_pct(row['حاشیه آخرین دوره'])}</td>
    <td style="padding:7px; border:1px solid #ddd;">{fmt_pct(row['نسبت مطالبات'])}</td>
</tr>
"""

html += "</tbody></table>"
st.markdown(html, unsafe_allow_html=True)

# راهنما
st.caption("""
**راهنما:**  
✅ **صادراتی** = شرکت دارای فروش صادراتی در سال مالی جاری  
📊 **پیشرفت** = درصدی از سال مالی که گزارش فروش برای آن ثبت شده است (مثلاً ۴ ماه = ۳۳٪)  
📈 **نسبت پوشش** = فروش محقق‌شده ÷ برآورد فروش سالانه  
✅ **جلوتر** = نسبت پوشش > پیشرفت (فروش از برنامه جلوتر است)  
⚠️ **عقب‌تر** = نسبت پوشش < پیشرفت (فروش از برنامه عقب‌تر است)  
📊 **طبق برنامه** = نسبت پوشش ≈ پیشرفت

🟢 **نسبت پوشش خوب** (> ۹۰%)  |  🟡 **متوسط** (۶۰-۹۰%)  |  🔴 **ضعیف** (< ۶۰%)  
P/E: 🟢 مناسب < ۵  |  🟡 متوسط ۵ تا ۷  |  🔴 بالا > ۷
""")


# ============================================================
# خروجی اکسل
# ============================================================
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def build_excel(df_export):
    wb = Workbook()
    ws = wb.active
    ws.title = "مقایسه شرکت‌ها"
    ws.sheet_view.rightToLeft = True

    headers = [
        "نماد", "نام", "صنعت", "صادراتی",
        "ارزش بازار", "پیشرفت سال مالی (%)", "تعداد ماه‌های گزارش",
        "برآورد فروش", "فروش محقق‌شده", "نسبت پوشش برآورد", "وضعیت",
        "میانگین حاشیه (%)", "برآورد سود", "سود تقسیمی",
        "P/E Forward", "P/D Forward", "P/S Forward",
        "حاشیه آخرین دوره (%)", "نسبت مطالبات (%)"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    yellow_fill = PatternFill("solid", fgColor="FEF9C3")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    alt_fill = PatternFill("solid", fgColor="F8FAFC")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin

    ws.row_dimensions[1].height = 28

    for r_idx, (_, row) in enumerate(df_export.iterrows(), 2):
        progress_val = row["پیشرفت سال مالی"]
        coverage_val = row["نسبت پوشش برآورد"]
        reported_months = row["تعداد ماه‌های گزارش"]
        has_export = row["صادراتی"]
        
        # وضعیت
        if coverage_val is not None and progress_val is not None:
            try:
                diff = float(coverage_val) - float(progress_val)
                if diff > 5:
                    status = "جلوتر"
                elif diff < -5:
                    status = "عقب‌تر"
                else:
                    status = "طبق برنامه"
            except:
                status = "—"
        else:
            status = "—"
        
        values = [
            row["نماد"], row["نام"], row["صنعت"], "بله" if has_export else "خیر",
            row["ارزش بازار"], progress_val, reported_months,
            row["برآورد فروش"], row["فروش محقق‌شده"], coverage_val, status,
            row["میانگین حاشیه"], row["برآورد سود"], row["برآورد سود تقسیمی"],
            row["P/E Forward"], row["P/D Forward"], row["P/S Forward"],
            row["حاشیه آخرین دوره"], row["نسبت مطالبات"],
        ]

        for c_idx, val in enumerate(values, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.alignment = cell_align
            cell.border = thin
            if r_idx % 2 == 0:
                cell.fill = alt_fill

            if c_idx in (1, 2, 3, 4, 11):
                cell.value = val if val is not None else "—"
                continue

            if c_idx == 6:  # پیشرفت سال مالی
                if val is None or pd.isna(val):
                    cell.value = "—"
                else:
                    cell.value = float(val)
                    cell.number_format = '0.00%'

            if c_idx == 7:  # تعداد ماه‌های گزارش
                cell.value = val if val is not None and val > 0 else "—"

            if c_idx in (5, 8, 9, 13, 14):
                if val is None or pd.isna(val):
                    cell.value = "—"
                else:
                    cell.value = float(val)
                    cell.number_format = '#,##0'

            elif c_idx == 10:  # نسبت پوشش برآورد
                if val is None or pd.isna(val):
                    cell.value = "—"
                else:
                    cell.value = float(val) / 100 if val > 1 else float(val)
                    cell.number_format = '0.00%'
                    if val >= 0.9:
                        cell.fill = green_fill
                    elif val >= 0.6:
                        cell.fill = yellow_fill
                    else:
                        cell.fill = red_fill

            elif c_idx in (12, 18, 19):
                if val is None or pd.isna(val):
                    cell.value = "—"
                else:
                    cell.value = float(val)
                    cell.number_format = '0.00%'

            elif c_idx in (15, 16, 17):
                if val is None or pd.isna(val):
                    cell.value = "—"
                else:
                    cell.value = float(val)
                    cell.number_format = '0.00'
                    if c_idx == 15:
                        pe = float(val)
                        if pe < 5:
                            cell.fill = green_fill
                        elif pe <= 7:
                            cell.fill = yellow_fill
                        else:
                            cell.fill = red_fill

    widths = [10, 22, 18, 12, 14, 16, 16, 14, 16, 14, 12, 14, 14, 12, 12, 12, 12, 16, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


try:
    excel_buffer = build_excel(df_view)
    st.download_button(
        label="📥 دانلود جدول مقایسه (Excel فرمت‌بندی‌شده)",
        data=excel_buffer,
        file_name="moghayese_sherkatha.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
except Exception as e:
    st.warning(f"خطا در ساخت اکسل: {e}")
    csv_df = df_view.copy()
    for col in ["میانگین حاشیه", "حاشیه آخرین دوره", "نسبت مطالبات", "نسبت پوشش برآورد", "پیشرفت سال مالی"]:
        csv_df[col] = csv_df[col].apply(lambda v: round(v * 100, 2) if v is not None and not pd.isna(v) else None)
    for col in ["P/E Forward", "P/D Forward", "P/S Forward"]:
        csv_df[col] = csv_df[col].apply(lambda v: round(v, 2) if v is not None and not pd.isna(v) else None)
    csv_buffer = BytesIO()
    csv_df.to_csv(csv_buffer, index=False, encoding="utf-8-sig")
    csv_buffer.seek(0)
    st.download_button(
        label="دانلود CSV",
        data=csv_buffer,
        file_name="moghayese_sherkatha.csv",
        mime="text/csv"
    )